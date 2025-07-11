# --- Imports principais
import os
import pandas as pd
import json
import pdfplumber
import customtkinter as ctk  # Importa customtkinter
import tkinter as tk # Mantido para messagebox e filedialog
from tkinter import filedialog, messagebox
from tkcalendar import Calendar, DateEntry # tkcalendar é necessário
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import win32com.client as win32 # pywin32 é necessário
import io
import traceback
import re
import datetime
import subprocess
import collections # Importado para usar o Counter
import threading # ### NOVO: Adicionado para tarefas em background
import pickle # ### NOVO: Para o sistema de caching

# --- Função Utilitária para Ler CSV Complexo (SEM ALTERAÇÃO) ---
def read_csv_with_encoding(file_path):
    encodings = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']
    common_delimiters = [',', ';', '\t']
    email_keywords = ['email', 'e-mail', 'usuário']
    entry_keywords = ['entrada', 'join', 'entrou']
    exit_keywords = ['saída', 'leave', 'saiu']
    end_marker_keyword = "compromisso da reunião"
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                lines = f.readlines()
            print(f"\nTentando ler CSV '{os.path.basename(file_path)}' Enc: '{encoding}'...")
            header_row_index = -1
            data_start_line = -1
            data_end_line = len(lines)
            print("  Procurando header CSV...")
            for i, line in enumerate(lines):
                lower_line = line.lower().strip()
                if not lower_line:
                    continue
                if any(kw in lower_line for kw in email_keywords) and \
                   any(kw in lower_line for kw in entry_keywords) and \
                   any(kw in lower_line for kw in exit_keywords):
                    print(f"  > Header CSV linha {i}.")
                    header_row_index = i
                    data_start_line = i + 1
                    break
            if header_row_index == -1:
                print(f"  <!> Header CSV não encontrado.")
                continue
            print(f"  Procurando fim CSV ('{end_marker_keyword}')...")
            found_end = False
            for i in range(data_start_line, len(lines)):
                if end_marker_keyword in lines[i].lower():
                    data_end_line = i
                    print(f"  > Fim CSV linha {i}.")
                    found_end = True
                    break
            if not found_end:
                print(f"  > Fim CSV não encontrado.")
            csv_content = "".join(lines[header_row_index:data_end_line])
            if not csv_content.strip():
                print(f"  <!> CSV vazio.")
                continue
            print(f"  Tentando pandas CSV (linhas {header_row_index}-{data_end_line-1})...")
            for delimiter in common_delimiters:
                try:
                    df = pd.read_csv(io.StringIO(csv_content), delimiter=delimiter, header=0, skipinitialspace=True, on_bad_lines='warn')
                    df.columns = df.columns.astype(str).str.strip()
                    print(f"    > Lido CSV Delim '{delimiter}'. Colunas: {list(df.columns)}")
                    df_cols_lower = [col.lower() for col in df.columns]
                    if any(any(kw in col for kw in email_keywords) for col in df_cols_lower) and \
                       any(any(kw in col for kw in entry_keywords) for col in df_cols_lower) and \
                       any(any(kw in col for kw in exit_keywords) for col in df_cols_lower):
                        print(f"  >> SUCESSO CSV!")
                        df.dropna(how='all', inplace=True)
                        return df
                    else:
                        print(f"    <!> Colunas essenciais CSV NÃO encontradas.")
                except Exception as e_pd:
                    print(f"    <!> Erro pandas CSV (Delim '{delimiter}'): {e_pd}")
            print(f"  <!> Nenhum delimitador CSV funcionou.")
        except UnicodeDecodeError:
            print(f"<!> Erro decode CSV '{encoding}'.")
        except FileNotFoundError:
            raise
        except Exception as e_file:
            print(f"<!> Erro GERAL CSV '{os.path.basename(file_path)}': {e_file}")
            traceback.print_exc()
    raise ValueError(f"Não foi possível ler CSV '{os.path.basename(file_path)}'.")

# --- Função Utilitária para Encontrar Coluna (SEM ALTERAÇÃO) ---
def find_col_ignore_case(df_columns, target_options):
    df_columns_lower = {str(col).lower().strip(): col for col in df_columns}
    for option in target_options:
        option_lower = option.lower()
        if option_lower in df_columns_lower:
            return df_columns_lower[option_lower]
    return None

# --- Classe de Login (UI com customtkinter) (SEM ALTERAÇÃO) ---
class LoginApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Login - Gerador de Relatórios")
        self.root.geometry("450x250")
        self.root.resizable(False, False)

        self.frame = ctk.CTkFrame(root, fg_color="transparent")
        self.frame.pack(expand=True, padx=30, pady=30)

        ctk.CTkLabel(self.frame, text="Bem-vindo!", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(0, 20))
        ctk.CTkLabel(self.frame, text="Seu Nome Completo:", font=ctk.CTkFont(size=14)).grid(row=1, column=0, padx=10, pady=10, sticky="e")
        
        self.nome_entry = ctk.CTkEntry(self.frame, width=220)
        self.nome_entry.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        self.nome_entry.focus()
        
        login_button = ctk.CTkButton(self.frame, text="Entrar", command=self.logar, width=120, font=ctk.CTkFont(size=12, weight="bold"))
        login_button.grid(row=2, column=0, columnspan=2, padx=10, pady=25)
        
        self.root.bind('<Return>', lambda event=None: self.logar())

    def logar(self):
        nome_completo = self.nome_entry.get().strip()
        if nome_completo:
            self.root.destroy()
            root_app = ctk.CTk() # Inicia a janela principal com customtkinter
            app = PresencaApp(root_app, nome_completo)
            root_app.mainloop()
        else:
            messagebox.showerror("Erro de Login", "Por favor, insira seu nome completo.", parent=self.root)

# --- Classe Principal da Aplicação (UI com customtkinter) ---
class PresencaApp:
    def __init__(self, root, nome_usuario):
        self.root = root
        self.nome_usuario = nome_usuario
        self.dates = []
        self.root.title(f"Gerador de Relatório de Presença - {self.nome_usuario}")
        self.root.geometry("950x780")
        self.root.minsize(900, 700)

        # ### NOVO: Flag para controlar a busca em background
        self.is_fetching_info = False

        # <<<<<<< NOVA FUNCIONALIDADE: Variáveis para a nova planilha e checkbox >>>>>>>
        self.master_spreadsheet_path = None
        self.master_professionals_path = None # Novo caminho para o relatório de profissionais
        self.use_professionals_report_var = tk.BooleanVar(value=False) # Variável para o checkbox
        # --- FIM DA NOVA FUNCIONALIDADE ---
        
        self._load_config()

        # Cores para botões específicos
        self.SUCCESS_COLOR = "#2ECC71"
        self.SUCCESS_HOVER_COLOR = "#27AE60"
        self.PREP_COLOR = "#E67E22"
        self.PREP_HOVER_COLOR = "#D35400"
        self.ERROR_COLOR = "#E74C3C"

        # --- Estrutura de Abas (CTkTabview) ---
        self.notebook = ctk.CTkTabview(root, height=750, width=900)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_config = self.notebook.add(" ⚙️ Configuração ")
        self.tab_arquivos = self.notebook.add(" 📄 Arquivos e Geração ")
        self.tab_emails_pos = self.notebook.add(" 📧 E-mails Pós-Treinamento ")

        self._create_tab_config()
        self._create_tab_arquivos()
        self._create_tab_emails_pos_treinamento()

        self._on_tipo_curso_change()
        self.add_instrutor_row()

    def _on_tipo_curso_change(self, *args):
        tipo = self.tipo_curso_var.get()
        hint = ""
        file_type_expected = " (XLSX, XLS ou CSV)."
        if tipo == "Digital":
            hint = "Presença Digital: XLSX c/ 'Data Entrada', 'Hora Entrada', 'Data Saída', 'Hora Saída', 'Email'."
            file_type_expected = " (XLSX ou XLS)."
        elif tipo == "Híbrido":
            hint = "Presença Híbrida: Combine formatos 'Digital' (XLSX) e/ou 'Presencial' (XLSX/CSV)."
        elif tipo == "Presencial":
            hint = "Presença Presencial: XLSX/CSV com coluna 'Email' ou 'Nome'."
        
        if hasattr(self, 'presenca_hint_label'):
            self.presenca_hint_label.configure(text=hint)
        
        if hasattr(self, 'presenca_text'):
            current_presenca_text = self.presenca_text.get("1.0", tk.END).strip()
            if current_presenca_text.startswith("Nenhum arquivo.") or not self.presenca_paths:
                self._update_file_text(self.presenca_text, None, f"Nenhum arquivo.{file_type_expected}")

    def _create_tab_config(self):
        scrollable_frame = ctk.CTkScrollableFrame(self.tab_config, label_text="Configuração do Treinamento", label_font=ctk.CTkFont(size=16, weight="bold"))
        scrollable_frame.pack(fill="both", expand=True, padx=5, pady=5)
        scrollable_frame.columnconfigure(0, weight=1)

        # --- Seção: Configuração da Planilha de Controle de Demandas ---
        master_sheet_frame = ctk.CTkFrame(scrollable_frame)
        master_sheet_frame.grid(row=0, column=0, pady=(10, 10), sticky="ew", padx=5) # Reduzido pady
        ctk.CTkLabel(master_sheet_frame, text="Conexão com Planilha de Controle de Demandas", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(10,5))
        info_label = "Para obter o caminho, sincronize a pasta do SharePoint, clique com o botão direito no arquivo e selecione 'Copiar como caminho'."
        ctk.CTkLabel(master_sheet_frame, text=info_label, font=ctk.CTkFont(size=11, slant="italic"), text_color="gray60", wraplength=500).pack(pady=(0,10), padx=20)
        ctk.CTkButton(master_sheet_frame, text="Definir/Alterar Planilha de Controle", command=self._set_master_spreadsheet).pack(pady=(0,10), padx=20)

        # <<<<<<< NOVA FUNCIONALIDADE: Seção para a nova planilha de Profissionais >>>>>>>
        professionals_sheet_frame = ctk.CTkFrame(scrollable_frame)
        professionals_sheet_frame.grid(row=1, column=0, pady=(0, 20), sticky="ew", padx=5)
        ctk.CTkLabel(professionals_sheet_frame, text="Conexão com Relatório de Profissionais (Base Opcional)", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(10,5))
        info_label_prof = "Use como alternativa à lista de convidados. O relatório final mostrará apenas os presentes."
        ctk.CTkLabel(professionals_sheet_frame, text=info_label_prof, font=ctk.CTkFont(size=11, slant="italic"), text_color="gray60", wraplength=500).pack(pady=(0,10), padx=20)
        ctk.CTkButton(professionals_sheet_frame, text="Definir/Alterar Relatório de Profissionais", command=self._set_master_professionals_spreadsheet).pack(pady=(0,10), padx=20)
        # --- FIM DA NOVA FUNCIONALIDADE ---

        # --- Seção: Detalhes do Treinamento ---
        details_frame = ctk.CTkFrame(scrollable_frame)
        details_frame.grid(row=2, column=0, pady=(0, 20), sticky="ew", padx=5) # Ajustado row para 2
        details_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(details_frame, text="Informações do Treinamento", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))

        self.labels = ["Número da Lista", "Nome do treinamento", "Turma", "Horário de Início", "Horário de Término",
                       "Horário de Saída do Intervalo (Opcional)", "Horário de Retorno do Intervalo (Opcional)",
                       "Carga Horária Bruta Total (h)", "Público", "CFC (Opcional)"]
        self.entries = {}
        grid_pady_inner = 6
        grid_padx_inner = (10, 10)

        for i, label_text in enumerate(self.labels):
            ctk.CTkLabel(details_frame, text=label_text + ":").grid(row=i+1, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
            entry = ctk.CTkEntry(details_frame, width=350)
            entry.grid(row=i+1, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
            self.entries[label_text] = entry
            
            if label_text == "Número da Lista":
                # ### ALTERADO: A função que ele chama agora é a nova função lançadora
                entry.bind("<FocusOut>", self._on_list_number_focus_out)

        # --- Seção dinâmica de instrutores ---
        instrutores_frame = ctk.CTkFrame(scrollable_frame)
        instrutores_frame.grid(row=3, column=0, pady=(0, 20), sticky="ew", padx=5) # Ajustado row para 3
        ctk.CTkLabel(instrutores_frame, text="Instrutores", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(10,15))

        self.instrutores_widgets = []
        self.instrutores_container = ctk.CTkFrame(instrutores_frame, fg_color="transparent")
        self.instrutores_container.pack(fill="x", expand=True, padx=10)
        ctk.CTkButton(instrutores_frame, text="+ Adicionar Instrutor", command=self.add_instrutor_row, fg_color=self.SUCCESS_COLOR, hover_color=self.SUCCESS_HOVER_COLOR).pack(pady=(10,10))
        
        # --- Seção: Tipo de Curso, Datas e Link ---
        config_dates_frame = ctk.CTkFrame(scrollable_frame)
        config_dates_frame.grid(row=4, column=0, pady=(0, 20), sticky="ew", padx=5) # Ajustado row para 4
        config_dates_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(config_dates_frame, text="Modalidade, Datas e Materiais", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))

        current_row_cf = 1
        ctk.CTkLabel(config_dates_frame, text="Tipo de Curso:").grid(row=current_row_cf, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
        self.tipo_curso_var = tk.StringVar(value="Digital")
        tipo_curso_options = ["Digital", "Presencial", "Híbrido"]
        self.tipo_curso_menu = ctk.CTkOptionMenu(config_dates_frame, variable=self.tipo_curso_var, values=tipo_curso_options, command=self._on_tipo_curso_change)
        self.tipo_curso_menu.grid(row=current_row_cf, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="w")
        current_row_cf += 1

        self.presenca_hint_label = ctk.CTkLabel(config_dates_frame, text="", text_color="gray60", font=ctk.CTkFont(slant="italic"))
        self.presenca_hint_label.grid(row=current_row_cf, column=1, padx=grid_padx_inner, pady=(0, grid_pady_inner+5), sticky="w")
        current_row_cf += 1

        ctk.CTkLabel(config_dates_frame, text="Link da Gravação:").grid(row=current_row_cf, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
        self.link_video_entry = ctk.CTkEntry(config_dates_frame, width=350)
        self.link_video_entry.grid(row=current_row_cf, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        current_row_cf += 1

        ctk.CTkLabel(config_dates_frame, text="Período/Data(s):").grid(row=current_row_cf, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
        date_frame = ctk.CTkFrame(config_dates_frame, fg_color="transparent")
        date_frame.grid(row=current_row_cf, column=1, sticky="w", padx=grid_padx_inner, pady=grid_pady_inner)
        self.date_entry = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.date_entry.pack(side=tk.LEFT)
        ctk.CTkButton(date_frame, text="Adicionar", command=self.add_date, width=80).pack(side=tk.LEFT, padx=(10, 5))
        ctk.CTkButton(date_frame, text="Limpar", command=self.clear_dates, width=80, fg_color="gray50", hover_color="gray40").pack(side=tk.LEFT, padx=(0, 0))
        current_row_cf += 1

        self.dates_text = ctk.CTkTextbox(config_dates_frame, height=70, width=350)
        self.dates_text.grid(row=current_row_cf, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        self.dates_text.insert(tk.END, "Nenhuma data adicionada.")
        self.dates_text.configure(state='disabled')

    def add_instrutor_row(self):
        row_frame = ctk.CTkFrame(self.instrutores_container, fg_color="transparent")
        row_frame.pack(fill="x", expand=True, pady=2)
        
        email_label = ctk.CTkLabel(row_frame, text="E-mail Instrutor:", width=110, anchor="e")
        email_label.pack(side="left", padx=(0, 5))
        
        email_entry = ctk.CTkEntry(row_frame, width=250)
        email_entry.pack(side="left", expand=True, fill="x", padx=5)

        ch_label = ctk.CTkLabel(row_frame, text="Carga Horária (HH:MM):", width=150, anchor="e")
        ch_label.pack(side="left", padx=(10, 5))
        
        ch_entry = ctk.CTkEntry(row_frame, width=70)
        ch_entry.pack(side="left", padx=5)

        remove_button = ctk.CTkButton(row_frame, text="✖", width=28, height=28, fg_color="transparent", border_width=1, border_color=self.ERROR_COLOR, text_color=self.ERROR_COLOR, hover_color=self.ERROR_COLOR,
                                   command=lambda rf=row_frame: self.remove_instrutor_row(rf))
        remove_button.pack(side="left", padx=10)

        self.instrutores_widgets.append((row_frame, email_entry, ch_entry))
        self._update_instrutor_buttons()

    def remove_instrutor_row(self, row_frame_to_remove):
        widgets_to_remove = None
        for widgets in self.instrutores_widgets:
            if widgets[0] == row_frame_to_remove:
                widgets_to_remove = widgets
                break
        
        if widgets_to_remove:
            widgets_to_remove[0].destroy()
            self.instrutores_widgets.remove(widgets_to_remove)
        
        self._update_instrutor_buttons()

    def _update_instrutor_buttons(self):
        is_visible = len(self.instrutores_widgets) > 1
        for i, (row_frame, _, _) in enumerate(self.instrutores_widgets):
            remove_button = row_frame.winfo_children()[-1]
            if is_visible:
                if not remove_button.winfo_ismapped():
                    remove_button.pack(side="left", padx=10)
            else:
                remove_button.pack_forget()

    # <<<<<<< NOVA FUNCIONALIDADE: Método para alternar a fonte dos convidados >>>>>>>
    def _toggle_convidados_source(self):
        use_professionals_report = self.use_professionals_report_var.get()

        if use_professionals_report:
            if not self.master_professionals_path or not os.path.exists(self.master_professionals_path):
                messagebox.showwarning("Configuração Necessária", "O caminho para o 'Relatório de Profissionais' não foi definido na aba de Configuração.", parent=self.root)
                self.use_professionals_report_var.set(False) # Desmarca o checkbox
                return

            # Desativa o anexo de convidados
            self.convidados_button.configure(state=tk.DISABLED)
            self.convidados_text.configure(state='normal')
            self.convidados_text.delete("1.0", tk.END)
            self.convidados_text.insert("1.0", "Usando o 'Relatório de Profissionais' como base.")
            self.convidados_text.configure(state='disabled')
        else:
            # Reativa o anexo de convidados
            self.convidados_button.configure(state=tk.NORMAL)
            self._update_file_text(self.convidados_text, self.convidados_path.get(), "Nenhum arquivo (XLSX, XLS, CSV).")
    # --- FIM DA NOVA FUNCIONALIDADE ---

    def _create_tab_arquivos(self):
        scrollable_frame = ctk.CTkScrollableFrame(self.tab_arquivos, label_text="Arquivos e Geração", label_font=ctk.CTkFont(size=16, weight="bold"))
        scrollable_frame.pack(fill="both", expand=True, padx=5, pady=5)
        scrollable_frame.columnconfigure(0, weight=1)

        # --- Seção: Ferramentas de Preparação ---
        prep_frame = ctk.CTkFrame(scrollable_frame)
        prep_frame.grid(row=0, column=0, pady=(10, 20), sticky="ew", padx=5)
        ctk.CTkLabel(prep_frame, text="Ferramentas de Preparação (Opcional)", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(10,15))
        
        btn_container = ctk.CTkFrame(prep_frame, fg_color="transparent")
        btn_container.pack(pady=(0,10), padx=10, fill="x")
        
        ctk.CTkButton(btn_container, text="Limpar CSV Presença (Teams)", command=self.limpar_csv_presenca, fg_color=self.PREP_COLOR, hover_color=self.PREP_HOVER_COLOR).pack(side="left", padx=(0,10), expand=True, fill="x")
        ctk.CTkButton(btn_container, text="Preparar Planilha (Acc. Journey)", command=self.preparar_accounting_journey, fg_color=self.PREP_COLOR, hover_color=self.PREP_HOVER_COLOR).pack(side="left", expand=True, fill="x")

        # --- Seção: Anexos ---
        anexos_frame = ctk.CTkFrame(scrollable_frame)
        anexos_frame.grid(row=1, column=0, pady=(0, 20), sticky="ew", padx=5)
        anexos_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(anexos_frame, text="Arquivos Necessários", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))
        
        self.convidados_path = tk.StringVar()
        self.presenca_paths = []
        self.nps_path = tk.StringVar()
        grid_pady_anexos = 8
        grid_padx_anexos = (10,10)
        current_row_anexos = 1

        # <<<<<<< NOVA FUNCIONALIDADE: Adicionado Checkbox e o botão é armazenado em self >>>>>>>
        self.convidados_button = ctk.CTkButton(anexos_frame, text="Anexar Lista de Convidados", command=self.upload_convidados, width=220)
        self.convidados_button.grid(row=current_row_anexos, column=0, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="e")
        self.convidados_text = ctk.CTkTextbox(anexos_frame, height=28)
        self.convidados_text.grid(row=current_row_anexos, column=1, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="we")
        self._update_file_text(self.convidados_text, None, "Nenhum arquivo (XLSX, XLS, CSV).")
        current_row_anexos += 1

        # Frame para o checkbox para alinhar corretamente
        checkbox_frame = ctk.CTkFrame(anexos_frame, fg_color="transparent")
        checkbox_frame.grid(row=current_row_anexos, column=1, padx=grid_padx_anexos, pady=(0, grid_pady_anexos), sticky="w")
        self.use_professionals_checkbox = ctk.CTkCheckBox(
            checkbox_frame, 
            text="Usar 'Relatório de Profissionais' como base (mostrar apenas presentes)", 
            variable=self.use_professionals_report_var,
            command=self._toggle_convidados_source
        )
        self.use_professionals_checkbox.pack(side="left")
        current_row_anexos += 1
        # --- FIM DA NOVA FUNCIONALIDADE ---

        ctk.CTkButton(anexos_frame, text="Anexar Lista(s) de Presença", command=self.upload_presenca, width=220).grid(row=current_row_anexos, column=0, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="e")
        self.presenca_text = ctk.CTkTextbox(anexos_frame, height=50)
        self.presenca_text.grid(row=current_row_anexos, column=1, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="we")
        self._update_file_text(self.presenca_text, None, "Nenhum arquivo. (O tipo ideal depende da Modalidade)")
        current_row_anexos += 1
        
        ctk.CTkButton(anexos_frame, text="Anexar NPS (PDF)", command=self.upload_nps, width=220).grid(row=current_row_anexos, column=0, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="e")
        self.nps_text = ctk.CTkTextbox(anexos_frame, height=28)
        self.nps_text.grid(row=current_row_anexos, column=1, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="we")
        self._update_file_text(self.nps_text, None, "Nenhum arquivo (PDF).")

        # --- Seção: Geração e Compartilhamento ---
        final_frame = ctk.CTkFrame(scrollable_frame)
        final_frame.grid(row=2, column=0, pady=(0, 20), sticky="ew", padx=5)
        final_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(final_frame, text="Geração e Compartilhamento", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))
        
        current_row_final = 1
        ctk.CTkLabel(final_frame, text="E-mails p/ compartilhar (separados por vírgula):").grid(row=current_row_final, column=0, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="e")
        self.emails_entry = ctk.CTkEntry(final_frame)
        self.emails_entry.grid(row=current_row_final, column=1, padx=grid_padx_anexos, pady=grid_pady_anexos, sticky="we")
        current_row_final += 1

        self.gerar_button = ctk.CTkButton(final_frame, text="🚀 Gerar Relatório de Presença", command=self.gerar_relatorio, font=ctk.CTkFont(size=14, weight="bold"), fg_color=self.SUCCESS_COLOR, hover_color=self.SUCCESS_HOVER_COLOR)
        self.gerar_button.grid(row=current_row_final, column=0, columnspan=2, pady=(20,10), padx=20)
        current_row_final += 1

        self.progress = ctk.CTkProgressBar(final_frame, orientation="horizontal")
        self.progress.grid(row=current_row_final, column=0, columnspan=2, pady=(0, 15), padx=20, sticky="ew")
        self.progress.set(0)

    def _create_tab_emails_pos_treinamento(self):
        # (Esta função não precisou de alterações)
        scrollable_frame = ctk.CTkScrollableFrame(self.tab_emails_pos, label_text="E-mails Pós-Treinamento", label_font=ctk.CTkFont(size=16, weight="bold"))
        scrollable_frame.pack(fill="both", expand=True, padx=5, pady=5)
        scrollable_frame.columnconfigure(0, weight=1)
        
        grid_pady_outer = (10, 15)
        grid_pady_inner = 8
        grid_padx_inner = (10, 10)
        text_height = 120

        presentes_frame = ctk.CTkFrame(scrollable_frame)
        presentes_frame.grid(row=0, column=0, pady=grid_pady_outer, sticky="ew")
        presentes_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(presentes_frame, text="✉️ E-mail para PRESENTES", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))

        ctk.CTkLabel(presentes_frame, text="Assunto (Presentes):").grid(row=1, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
        self.assunto_presentes_entry = ctk.CTkEntry(presentes_frame)
        self.assunto_presentes_entry.grid(row=1, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        self.assunto_presentes_entry.insert(0, "Certificado e Agradecimento - Treinamento {{nome_treinamento}}")

        ctk.CTkLabel(presentes_frame, text="Corpo (Presentes):\n(Use {{nome}}, {{nome_treinamento}})").grid(row=2, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="ne")
        self.corpo_presentes_text = ctk.CTkTextbox(presentes_frame, height=text_height, wrap=tk.WORD)
        self.corpo_presentes_text.grid(row=2, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        self.corpo_presentes_text.insert(tk.END, "Olá {{nome}},\n\nObrigado por participar do treinamento \"{{nome_treinamento}}\"!\n\nSeu certificado (se aplicável) será enviado em breve ou segue anexo.\n\nAtenciosamente,\nA Equipe")

        faltantes_frame = ctk.CTkFrame(scrollable_frame)
        faltantes_frame.grid(row=1, column=0, pady=grid_pady_outer, sticky="ew")
        faltantes_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(faltantes_frame, text="✉️ E-mail para FALTANTES", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))

        ctk.CTkLabel(faltantes_frame, text="Assunto (Faltantes):").grid(row=1, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
        self.assunto_faltantes_entry = ctk.CTkEntry(faltantes_frame)
        self.assunto_faltantes_entry.grid(row=1, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        self.assunto_faltantes_entry.insert(0, "Ausência no Treinamento {{nome_treinamento}}")

        ctk.CTkLabel(faltantes_frame, text="Corpo (Faltantes):\n(Use {{nome}}, {{nome_treinamento}})").grid(row=2, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="ne")
        self.corpo_faltantes_text = ctk.CTkTextbox(faltantes_frame, height=text_height, wrap=tk.WORD)
        self.corpo_faltantes_text.grid(row=2, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        self.corpo_faltantes_text.insert(tk.END, "Olá {{nome}},\n\nNotamos sua ausência no treinamento \"{{nome_treinamento}}\".\n\nCaso tenha ocorrido algum imprevisto ou deseje informações sobre futuras turmas, por favor, entre em contato.\n\nAtenciosamente,\nA Equipe")

        anexos_email_frame = ctk.CTkFrame(scrollable_frame)
        anexos_email_frame.grid(row=2, column=0, pady=grid_pady_outer, sticky="ew")
        anexos_email_frame.columnconfigure(1, weight=1)
        ctk.CTkLabel(anexos_email_frame, text="📎 Anexos Comuns (Opcional)", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=(10,15))

        ctk.CTkButton(anexos_email_frame, text="Anexar Arquivos", command=self._upload_anexos_pos_treinamento, width=150).grid(row=1, column=0, padx=grid_padx_inner, pady=grid_pady_inner, sticky="e")
        self.anexos_pos_treinamento_text = ctk.CTkTextbox(anexos_email_frame, height=50)
        self.anexos_pos_treinamento_text.grid(row=1, column=1, padx=grid_padx_inner, pady=grid_pady_inner, sticky="we")
        self._update_file_text(self.anexos_pos_treinamento_text, None, "Nenhum arquivo anexado.")
        self.anexos_pos_treinamento_paths = []

        action_frame_emails = ctk.CTkFrame(scrollable_frame, fg_color="transparent")
        action_frame_emails.grid(row=3, column=0, pady=(10,0), sticky="ew")
        action_frame_emails.columnconfigure(0, weight=1)

        self.enviar_pos_treinamento_button = ctk.CTkButton(action_frame_emails, text="🚀 Enviar E-mails Pós-Treinamento", command=self._enviar_emails_pos_treinamento_action, font=ctk.CTkFont(size=14, weight="bold"), fg_color=self.SUCCESS_COLOR, hover_color=self.SUCCESS_HOVER_COLOR)
        self.enviar_pos_treinamento_button.grid(row=0, column=0, pady=(15,5), padx=20)

        self.progress_emails = ctk.CTkProgressBar(action_frame_emails, orientation="horizontal")
        self.progress_emails.grid(row=1, column=0, pady=(0,10), padx=20, sticky="ew")
        self.progress_emails.set(0)

        log_email_frame = ctk.CTkFrame(scrollable_frame)
        log_email_frame.grid(row=4, column=0, pady=grid_pady_outer, sticky="nsew")
        log_email_frame.columnconfigure(0, weight=1)
        log_email_frame.rowconfigure(1, weight=1)
        ctk.CTkLabel(log_email_frame, text="Log de Envio", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, pady=(10,5))
        
        self.log_envio_emails_text = ctk.CTkTextbox(log_email_frame, wrap=tk.WORD, height=100)
        self.log_envio_emails_text.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0,10))
        self.log_envio_emails_text.configure(state=tk.DISABLED)

    # --- MÉTODOS DE BACKEND E LÓGICA ---

    def _get_instrutores_data(self):
        instrutores = []
        emails = []
        for _, email_entry, ch_entry in self.instrutores_widgets:
            email = email_entry.get().strip()
            ch = ch_entry.get().strip()
            if email and ch:
                if not re.fullmatch(r"(\d{1,3}):(\d{2})", ch):
                    raise ValueError(f"Formato de Carga Horária inválido para o instrutor '{email}'. Use HH:MM.")
                instrutores.append({'email': email, 'ch': ch})
                emails.append(email.lower())
            elif email or ch:
                 raise ValueError("Preencha tanto o e-mail quanto a carga horária para cada instrutor.")
        return instrutores, emails

    def _log_message(self, message):
        print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}")

    def _load_config(self):
        try:
            with open("config.json", "r") as f:
                config = json.load(f)
                
                # Carrega planilha de controle
                path_control = config.get("master_spreadsheet_path")
                if path_control and os.path.exists(path_control):
                    self.master_spreadsheet_path = path_control
                    self._log_message(f"Planilha de controle carregada: {os.path.basename(path_control)}")
                else:
                    self._log_message("Caminho da planilha de controle no config.json é inválido.")
                
                # <<<<<<< NOVA FUNCIONALIDADE: Carrega a nova planilha de profissionais >>>>>>>
                path_prof = config.get("master_professionals_path")
                if path_prof and os.path.exists(path_prof):
                    self.master_professionals_path = path_prof
                    self._log_message(f"Relatório de profissionais carregado: {os.path.basename(path_prof)}")
                else:
                    self._log_message("Caminho do relatório de profissionais no config.json é inválido.")
                # --- FIM DA NOVA FUNCIONALIDADE ---

        except (FileNotFoundError, json.JSONDecodeError):
            self._log_message("Arquivo de configuração (config.json) não encontrado ou inválido.")

    def _save_config(self):
        # <<<<<<< NOVA FUNCIONALIDADE: Salva ambos os caminhos >>>>>>>
        config_data = {
            "master_spreadsheet_path": self.master_spreadsheet_path,
            "master_professionals_path": self.master_professionals_path
        }
        with open("config.json", "w") as f:
            json.dump(config_data, f, indent=4)
        self._log_message("Configuração salva.")
        # --- FIM DA NOVA FUNCIONALIDADE ---

    def _set_master_spreadsheet(self):
        path = filedialog.askopenfilename(
            title="Selecione sua Planilha de Controle (do seu diretório local)",
            filetypes=[("Planilhas Excel", "*.xlsx;*.xls")],
            parent=self.root
        )
        if path:
            self.master_spreadsheet_path = path
            self._save_config()
            messagebox.showinfo("Sucesso", f"Caminho da planilha de controle definido para:\n{path}", parent=self.root)

    # <<<<<<< NOVA FUNCIONALIDADE: Método para definir o caminho da nova planilha >>>>>>>
    def _set_master_professionals_spreadsheet(self):
        path = filedialog.askopenfilename(
            title="Selecione seu Relatório de Profissionais (do seu diretório local)",
            filetypes=[("Planilhas Excel", "*.xlsx;*.xls")],
            parent=self.root
        )
        if path:
            self.master_professionals_path = path
            self._save_config()
            messagebox.showinfo("Sucesso", f"Caminho do Relatório de Profissionais definido para:\n{path}", parent=self.root)
            # Se o checkbox estiver marcado, atualizamos o estado da UI
            if self.use_professionals_report_var.get():
                self._toggle_convidados_source()
    # --- FIM DA NOVA FUNCIONALIDADE ---
    
    # ### ALTERADO: Esta função agora inicia a thread em background
    def _on_list_number_focus_out(self, event=None):
        """
        Inicia a busca de informações do treinamento em uma thread separada 
        para não travar a interface.
        """
        if self.is_fetching_info:
            self._log_message("Busca de informações já em andamento. Aguarde.")
            return

        list_number = self.entries["Número da Lista"].get().strip()
        if list_number:
            self.is_fetching_info = True
            # Fornece um feedback visual para o usuário
            self.entries["Número da Lista"].configure(state="disabled")
            self.entries["Nome do treinamento"].delete(0, tk.END)
            self.entries["Nome do treinamento"].insert(0, "Buscando informações, aguarde...")
            
            # Cria e inicia a thread que fará o trabalho pesado
            thread = threading.Thread(
                target=self._fetch_training_info_worker, 
                args=(list_number,),
                daemon=True # Permite que a aplicação feche mesmo se a thread estiver rodando
            )
            thread.start()

    # ### ALTERADO: A lógica de busca foi movida para uma função separada e agora usa cache.
    def _fetch_training_info_worker(self, list_number):
        """
        Busca informações do treinamento, utilizando um sistema de cache para acelerar leituras repetidas.
        """
        try:
            # Pega o DataFrame principal, usando o cache se possível.
            df_master = self._get_master_df_with_caching()
            if df_master is None: # Ocorreu um erro na leitura/cache, a notificação já foi enviada.
                return 

            COL_NAMES = {
                'numero_lista': 'Numero de Lista', 'nome': 'Nome do Treinamento', 'publico': 'Público Alvo',
                'horario': 'Horário', 'instrutor': 'Instrutor(a)', 'formato': 'Formato'
            }
            
            search_col_name = COL_NAMES['numero_lista']
            # A busca no DataFrame já carregado é muito rápida
            training_data = df_master[df_master[search_col_name].astype(str).str.strip() == str(list_number)]

            if training_data.empty:
                error_info = ("Não Encontrado", f"Nenhum treinamento encontrado com o Número da Lista '{list_number}'.")
                self.root.after(0, self._show_fetch_error, error_info)
                return

            data_row = training_data.iloc[0]
            self._log_message(f"Dados encontrados para a LP nº {list_number}. Preparando para atualizar UI...")
            
            def get_data_by_col_name(row, col_key):
                col_name = COL_NAMES[col_key]
                return str(row[col_name]) if col_name in row and pd.notna(row[col_name]) else ""

            horario_str = get_data_by_col_name(data_row, 'horario')
            match = re.search(r'(\d{1,2}:\d{2}).*?(\d{1,2}:\d{2})', horario_str)
            inicio, termino = match.groups() if match else ("", "")

            formato_from_sheet = get_data_by_col_name(data_row, 'formato').lower().strip()
            ui_course_type = "Digital"
            if formato_from_sheet == 'presencial': ui_course_type = "Presencial"
            elif formato_from_sheet == 'hibrido': ui_course_type = "Híbrido"

            results = {
                "Nome do treinamento": get_data_by_col_name(data_row, 'nome'),
                "Público": get_data_by_col_name(data_row, 'publico'),
                "Horário de Início": inicio, "Horário de Término": termino,
                "instrutor_email": get_data_by_col_name(data_row, 'instrutor'),
                "tipo_curso": ui_course_type
            }
            
            self.root.after(0, self._update_ui_with_fetched_data, results)

        except Exception as e:
            error_info = ("Erro na Busca", f"Ocorreu um erro ao buscar os dados:\n{e}")
            self.root.after(0, self._show_fetch_error, error_info)
            traceback.print_exc()
        finally:
            self.root.after(0, self._finalize_fetch)

    # ### NOVO: Função de caching para a planilha mestre
    def _get_master_df_with_caching(self):
        """
        Carrega o DataFrame da planilha de controle, usando um arquivo de cache (.pkl)
        para evitar a releitura lenta do Excel se o arquivo não mudou.
        """
        if not self.master_spreadsheet_path or not os.path.exists(self.master_spreadsheet_path):
            error_info = ("Configuração Necessária", "O caminho para a Planilha de Controle não foi definido ou é inválido.")
            self.root.after(0, self._show_fetch_error, error_info)
            return None

        # Define um nome para o arquivo de cache
        cache_file = "master_schedule_cache.pkl"

        try:
            cache_is_valid = False
            if os.path.exists(cache_file):
                # Compara a data de modificação do cache com a do arquivo original
                cache_mod_time = os.path.getmtime(cache_file)
                excel_mod_time = os.path.getmtime(self.master_spreadsheet_path)
                if cache_mod_time > excel_mod_time:
                    cache_is_valid = True

            if cache_is_valid:
                self._log_message("Cache válido encontrado! Carregando dados da planilha a partir do cache (rápido).")
                with open(cache_file, 'rb') as f:
                    df_master = pickle.load(f)
                return df_master
            
            # Se o cache não for válido, fazemos a leitura lenta
            self._log_message("Cache inválido ou inexistente. Lendo o arquivo Excel completo (lento)...")
            TARGET_SHEET_NAME = "Programação Geral 2025"
            COL_NAMES = {
                'numero_lista': 'Numero de Lista', 'nome': 'Nome do Treinamento', 'publico': 'Público Alvo',
                'horario': 'Horário', 'instrutor': 'Instrutor(a)', 'formato': 'Formato'
            }
            
            df_temp = pd.read_excel(self.master_spreadsheet_path, sheet_name=TARGET_SHEET_NAME, header=None, engine='openpyxl')

            header_row_index = -1
            keywords_to_find = list(COL_NAMES.values())
            for i, row in df_temp.head(20).iterrows():
                row_values = [str(cell).strip() for cell in row.dropna()]
                if len([kw for kw in keywords_to_find if kw in row_values]) >= 3:
                    header_row_index = i
                    break
            
            if header_row_index == -1:
                raise ValueError(f"Não foi possível encontrar a linha de cabeçalho na aba '{TARGET_SHEET_NAME}'.")

            df_master = df_temp.copy()
            df_master.columns = df_master.iloc[header_row_index]
            df_master = df_master.drop(df_master.index[0:header_row_index + 1])
            df_master.columns = df_master.columns.str.strip()
            
            # Salva o DataFrame processado no cache para a próxima vez
            self._log_message(f"Salvando {len(df_master)} linhas no arquivo de cache '{cache_file}' para uso futuro.")
            with open(cache_file, 'wb') as f:
                pickle.dump(df_master, f)
                
            return df_master

        except Exception as e:
            # Se der erro, garante que a thread principal seja notificada
            error_title = "Erro ao Processar Planilha"
            if "Worksheet named" in str(e):
                error_title = "Erro de Aba"
                e = f"A aba 'Programação Geral 2025' não foi encontrada."
            
            self.root.after(0, self._show_fetch_error, (error_title, str(e)))
            traceback.print_exc()
            return None

    # ### NOVO: Função para atualizar a UI com os dados, de forma segura (NÃO MUDA NESTA ETAPA)
    def _update_ui_with_fetched_data(self, data):
        """
        Atualiza os widgets da interface com os dados recebidos da thread worker.
        Esta função SEMPRE executa na thread principal.
        """
        self.entries["Nome do treinamento"].delete(0, tk.END)
        self.entries["Nome do treinamento"].insert(0, data.get("Nome do treinamento", ""))
        
        self.entries["Público"].delete(0, tk.END)
        self.entries["Público"].insert(0, data.get("Público", ""))
        
        self.entries["Horário de Início"].delete(0, tk.END)
        self.entries["Horário de Início"].insert(0, data.get("Horário de Início", ""))

        self.entries["Horário de Término"].delete(0, tk.END)
        self.entries["Horário de Término"].insert(0, data.get("Horário de Término", ""))
        
        if data.get("instrutor_email") and self.instrutores_widgets:
            self.instrutores_widgets[0][1].delete(0, tk.END)
            self.instrutores_widgets[0][1].insert(0, data["instrutor_email"])
            
        self.tipo_curso_var.set(data.get("tipo_curso", "Digital"))
        self._log_message("Campos da interface atualizados com sucesso.")

    # ### NOVO: Função para mostrar erros na UI, de forma segura (NÃO MUDA NESTA ETAPA)
    def _show_fetch_error(self, error_info):
        """Mostra uma messagebox de erro. Executa na thread principal."""
        title, message = error_info
        # Limpa o campo de "buscando..."
        if self.entries["Nome do treinamento"].get() == "Buscando informações, aguarde...":
             self.entries["Nome do treinamento"].delete(0, tk.END)
        messagebox.showerror(title, message, parent=self.root)

    # ### NOVO: Função para reativar a UI após a busca (sucesso ou falha) (NÃO MUDA NESTA ETAPA)
    def _finalize_fetch(self):
        """Reativa a UI após a conclusão da busca em background."""
        self.is_fetching_info = False
        self.entries["Número da Lista"].configure(state="normal")
        # Se o campo de nome ainda estiver com a mensagem de "buscando", limpa ele.
        if "Buscando informações" in self.entries["Nome do treinamento"].get():
            self.entries["Nome do treinamento"].delete(0, tk.END)
        self._log_message("Busca de informações finalizada.")
            
    def _parse_hh_mm_to_hours(self, time_str: str) -> float:
        if not time_str:
            raise ValueError("O campo de Carga Horária não pode estar vazio.")
        
        time_str = str(time_str).strip()
        match = re.fullmatch(r"(\d{1,3}):(\d{2})", time_str)
        if not match:
            try:
                return float(str(time_str).replace(',', '.'))
            except ValueError:
                 raise ValueError(f"Formato de Carga Horária '{time_str}' inválido. Use HH:MM (ex: 08:00) ou um número (ex: 8.5).")
        
        try:
            hours = int(match.group(1))
            minutes = int(match.group(2))
        except ValueError:
            raise ValueError("Horas e minutos devem ser numéricos.")

        if not (0 <= minutes < 60):
            raise ValueError("Minutos devem estar entre 00 e 59.")
        if not (0 <= hours < 1000):
            raise ValueError("Valor de horas inválido (deve ser entre 0 e 999).")
            
        total_hours = hours + (minutes / 60.0)
        if total_hours <= 0:
            raise ValueError("A carga horária deve ser um valor positivo.")
        return total_hours

    def _parse_time_to_total_minutes(self, time_str: str) -> int:
        if not time_str:
            raise ValueError("Horário (HH:MM) não pode ser vazio.")

        match = re.fullmatch(r"(\d{1,2}):(\d{2})", time_str)
        if not match:
            raise ValueError(f"Formato de horário inválido '{time_str}'. Use HH:MM (ex: 12:00).")

        try:
            hours = int(match.group(1))
            minutes = int(match.group(2))
        except ValueError:
            raise ValueError(f"Horas e minutos em '{time_str}' devem ser numéricos.")

        if not (0 <= hours <= 23):
            raise ValueError(f"Hora '{hours}' em '{time_str}' inválida. Deve estar entre 00 e 23.")
        if not (0 <= minutes <= 59):
            raise ValueError(f"Minutos '{minutes}' em '{time_str}' inválidos. Devem estar entre 00 e 59.")

        return hours * 60 + minutes

    def _get_effective_training_hours(self):
        carga_horaria_total_str = self.entries["Carga Horária Bruta Total (h)"].get().strip()
        total_gross_hours = self._parse_hh_mm_to_hours(carga_horaria_total_str)

        intervalo_saida_str = self.entries["Horário de Saída do Intervalo (Opcional)"].get().strip()
        intervalo_retorno_str = self.entries["Horário de Retorno do Intervalo (Opcional)"].get().strip()

        num_dias = len(self.presenca_paths)
        if num_dias == 0:
            self._log_message("AVISO: Nenhuma lista de presença encontrada para calcular totais. Assumindo 1 dia.")
            num_dias = 1
            
        daily_gross_hours = total_gross_hours / num_dias
        
        daily_interval_hours = 0.0
        if intervalo_saida_str and intervalo_retorno_str:
            saida_minutos = self._parse_time_to_total_minutes(intervalo_saida_str)
            retorno_minutos = self._parse_time_to_total_minutes(intervalo_retorno_str)
            if retorno_minutos <= saida_minutos:
                raise ValueError("O horário de retorno do intervalo deve ser após o horário de saída.")
            daily_interval_hours = (retorno_minutos - saida_minutos) / 60.0
        elif intervalo_saida_str or intervalo_retorno_str:
            raise ValueError("Para considerar o intervalo, preencha ambos os horários (saída e retorno) ou deixe ambos vazios.")
        
        total_interval_hours = daily_interval_hours * num_dias
        
        net_hours = total_gross_hours - total_interval_hours
        
        self._log_message(f"Cálculo de Carga Horária para Percentual:")
        self._log_message(f"  - C.H. Bruta Total (Input): {total_gross_hours:.2f}h")
        self._log_message(f"  - Número de Dias (Listas): {num_dias}")
        self._log_message(f"  - C.H. Bruta por Dia (Calculado): {daily_gross_hours:.2f}h")
        self._log_message(f"  - Desconto Intervalo por Dia: {daily_interval_hours:.2f}h")
        self._log_message(f"  - Desconto Intervalo Total: {total_interval_hours:.2f}h")
        self._log_message(f"  - C.H. LÍQUIDA FINAL (Base 100%): {net_hours:.2f}h")

        if net_hours <= 0:
            raise ValueError(f"A carga horária líquida final ({net_hours:.2f}h) é zero ou negativa. Verifique os valores inseridos.")
        
        return net_hours, total_gross_hours, total_interval_hours, daily_gross_hours, daily_interval_hours

    def add_date(self):
        try:
            date_obj = self.date_entry.get_date()
            if date_obj not in self.dates:
                self.dates.append(date_obj)
                self.dates.sort()
                self.dates_text.configure(state='normal')
                self.dates_text.delete(1.0, tk.END)
                if self.dates:
                    self.dates_text.insert(tk.END, "\n".join([d.strftime("%d/%m/%Y") for d in self.dates]))
                else:
                    self.dates_text.insert(tk.END, "Nenhuma data adicionada.")
                self.dates_text.configure(state='disabled')
            else:
                messagebox.showinfo("Data Duplicada", "Esta data já foi adicionada.", parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro de Data", f"Não foi possível adicionar a data: {e}", parent=self.root)
            
    def clear_dates(self):
        if not self.dates:
            messagebox.showinfo("Limpar Datas", "Nenhuma data para limpar.", parent=self.root)
            return
        if messagebox.askyesno("Confirmar Limpeza", "Tem certeza que deseja limpar todas as datas adicionadas?", parent=self.root):
            self.dates.clear()
            self.dates_text.configure(state='normal')
            self.dates_text.delete(1.0, tk.END)
            self.dates_text.insert(tk.END, "Nenhuma data adicionada.")
            self.dates_text.configure(state='disabled')
            messagebox.showinfo("Datas Limpas", "Todas as datas foram removidas.", parent=self.root)            

    def _update_file_text(self, text_widget, file_path_or_list, placeholder="Nenhum arquivo selecionado."):
        text_widget.configure(state='normal')
        text_widget.delete("1.0", tk.END)
        display_text = placeholder
        if isinstance(file_path_or_list, list) and file_path_or_list:
            display_text = "\n".join([os.path.basename(p) for p in file_path_or_list])
        elif isinstance(file_path_or_list, str) and file_path_or_list:
            display_text = os.path.basename(file_path_or_list)
        text_widget.insert("1.0", display_text)
        text_widget.configure(state='disabled')

    def upload_convidados(self):
        file_path = filedialog.askopenfilename(title="Selecionar Lista de Convidados (XLSX, XLS, CSV)", filetypes=[("Planilhas", "*.xlsx;*.xls;*.csv"), ("Todos", "*.*")], parent=self.root)
        if file_path:
            self.convidados_path.set(file_path)
            self._update_file_text(self.convidados_text, file_path)

    def upload_presenca(self):
        tipo = self.tipo_curso_var.get()
        title = "Selecionar Lista(s) de Presença"
        filetypes = [("Planilhas e CSV", "*.xlsx;*.xls;*.csv"), ("Todos", "*.*")]
        placeholder_suffix = " (XLSX, XLS ou CSV)."
        if tipo == "Digital":
            title += " (Excel c/ Data/Hora)"
            filetypes = [("Excel", "*.xlsx;*.xls"), ("Todos", "*.*")]
            placeholder_suffix = " (XLSX ou XLS)."
        elif tipo == "Presencial":
            title += " (Excel/CSV c/ ID)"
        elif tipo == "Híbrido":
            title += " (Excel/CSV)"

        file_paths = filedialog.askopenfilenames(title=title, filetypes=filetypes, parent=self.root)
        if file_paths:
            valid_paths = list(file_paths)
            if tipo == "Digital":
                invalid_ext = [os.path.basename(p) for p in file_paths if not p.lower().endswith(('.xlsx', '.xls'))]
                if invalid_ext:
                    messagebox.showwarning("Extensão Inesperada", f"Para '{tipo}', arquivos .xlsx/.xls são esperados.\nOutras extensões podem não funcionar:\n" + "\n".join(invalid_ext), parent=self.root)
            self.presenca_paths = valid_paths
            if self.presenca_paths:
                self._update_file_text(self.presenca_text, self.presenca_paths)
            else:
                self._update_file_text(self.presenca_text, None, f"Nenhuma lista válida anexada.{placeholder_suffix}")
        self._on_tipo_curso_change()

    def _upload_anexos_pos_treinamento(self):
        file_paths = filedialog.askopenfilenames(title="Selecionar Arquivos para Anexar aos E-mails", parent=self.root)
        if file_paths:
            self.anexos_pos_treinamento_paths = list(file_paths)
            self._update_file_text(self.anexos_pos_treinamento_text, self.anexos_pos_treinamento_paths)
        else:
            self.anexos_pos_treinamento_paths = []
            self._update_file_text(self.anexos_pos_treinamento_text, None, "Nenhum arquivo anexado.")

    def upload_nps(self):
        file_path = filedialog.askopenfilename(title="Selecionar Arquivo NPS (PDF)", filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")], parent=self.root)
        if file_path:
            self.nps_path.set(file_path)
            self._update_file_text(self.nps_text, file_path)
    
    def limpar_csv_presenca(self):
        print("\n--- Iniciando Limpeza de CSV de Presença ---")
        csv_separator = ';'
        csv_encodings_to_try = ['utf-16', 'utf-8-sig', 'utf-8']

        lang_settings = {
            "pt": {
                "start_marker": "3. Atividades em Reunião",
                "end_marker": "4. Compromisso da Reunião",
                "datetime_cols_map": {"Horário de Entrada": "Entrada", "Horário de Saída": "Saída"},
                "rename_to_pt_map": {"Nome": "Nome", "Duração": "Duração", "Email": "Email", "Função": "Função"}
            },
            "en": {
                "start_marker": "3. In-Meeting Activities",
                "end_marker": "4. Meeting Engagement",
                "datetime_cols_map": {"Join Time": "Entrada", "Leave Time": "Saída"},
                "rename_to_pt_map": {"Name": "Nome", "Duration": "Duração", "Email": "Email", "Role": "Função"}
            }
        }
        selected_lang_config = None
        data_delimiter = '\t'
        date_time_separator = ', '

        input_csv_path = filedialog.askopenfilename(
            title="Selecione o CSV de Presença Bruto para Limpar",
            filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")],
            parent=self.root
        )
        if not input_csv_path:
            print("   > Nenhuma arquivo selecionado.")
            messagebox.showinfo("Cancelado", "Nenhum arquivo CSV selecionado.", parent=self.root)
            return

        input_filename = os.path.basename(input_csv_path)
        output_folder = os.path.dirname(input_csv_path)
        base_filename_out = os.path.splitext(input_filename)[0]
        output_xlsx_path = os.path.join(output_folder, f"{base_filename_out}_datas_separadas.xlsx")
        print(f"   > Arquivo selecionado: {input_filename}")
        print(f"   > Saída será: {os.path.basename(output_xlsx_path)}")

        try:
            df_full = None
            used_encoding = None
            for enc in csv_encodings_to_try:
                try:
                    print(f"   > Tentando ler CSV (sep='{csv_separator}', enc='{enc}')...")
                    df_full = pd.read_csv(input_csv_path, header=None, sep=csv_separator, encoding=enc, dtype='string', skip_blank_lines=False)
                    used_encoding = enc
                    print(f"      > Leitura com '{enc}' bem-sucedida.")
                    break 
                except UnicodeDecodeError:
                    print(f"      <!> Falha ao decodificar com '{enc}'.")
                except pd.errors.ParserError as pe:
                    print(f"      <!> Erro de parser com '{enc}': {pe}. Verifique o separador.")
                except FileNotFoundError:
                    messagebox.showerror("Erro", f"Arquivo não encontrado:\n{input_csv_path}", parent=self.root)
                    return
                except Exception as e_read:
                    print(f"      <!> Erro inesperado ao ler com '{enc}': {e_read}")

            if df_full is None:
                messagebox.showerror("Erro Leitura CSV", f"Não foi possível ler o arquivo CSV '{input_filename}' com as codificações testadas ({', '.join(csv_encodings_to_try)}).\nVerifique a codificação e o separador ('{csv_separator}') do arquivo.", parent=self.root)
                return

            print(f"   > Leitura OK: {len(df_full)} linhas.")
            print("   > Procurando marcadores...")
            start_row_index = None
            end_row_index = None
            final_cleaned_df = None
            
            first_col_str_series = df_full.iloc[:, 0].astype(str)
            for lang_code, config in lang_settings.items():
                if first_col_str_series.str.contains(config["start_marker"], na=False, regex=False).any():
                    selected_lang_config = config
                    print(f"      -> Marcador de início '{config['start_marker']}' (idioma: {lang_code.upper()}) encontrado.")
                    break
            
            if not selected_lang_config:
                all_starts = [cfg["start_marker"] for cfg in lang_settings.values()]
                raise ValueError(f"Nenhum marcador de início conhecido ({', '.join(all_starts)}) encontrado na primeira coluna.")

            current_start_marker = selected_lang_config["start_marker"]
            current_end_marker = selected_lang_config["end_marker"]
            start_matches = df_full[first_col_str_series.str.contains(current_start_marker, na=False, regex=False)]

            if not start_matches.empty:
                start_row_index = start_matches.index[0]
                print(f"      -> Início: índice {start_row_index} (linha {start_row_index + 1})")
            else:
                raise ValueError(f"Marcador de início '{current_start_marker}' não encontrado (após seleção de idioma).")

            df_search_area = df_full.iloc[start_row_index + 1:]
            end_matches = df_search_area[df_search_area.iloc[:, 0].astype(str).str.contains(current_end_marker, na=False, regex=False)]
            if not end_matches.empty:
                end_row_index = end_matches.index[0]
                print(f"      -> Fim: índice {end_row_index} (linha {end_row_index + 1})")
            else:
                end_row_index = len(df_full)
                print(f"      -> AVISO: Marcador de fim '{current_end_marker}' não encontrado após o início. Usando o final do arquivo.")
            
            header_row_index = start_row_index + 1
            data_start_index = header_row_index + 1
            
            if header_row_index < end_row_index and data_start_index <= end_row_index and header_row_index < len(df_full):
                header_string = df_full.iloc[header_row_index, 0]
                actual_headers = []
                if pd.notna(header_string):
                    actual_headers = [h.strip() for h in header_string.split(data_delimiter)]
                    print(f"      -> Cabeçalhos extraídos: {actual_headers}")
                    if not actual_headers:
                        raise ValueError("Cabeçalho vazio após split.")
                else:
                    raise ValueError(f"Linha do cabeçalho (índice {header_row_index}) parece vazia.")
                
                extracted_section_series = df_full.iloc[data_start_index : end_row_index, 0].copy()
                if not extracted_section_series.empty and len(actual_headers) > 0:
                    print(f"      -> Dividindo {len(extracted_section_series)} linhas de dados por '{repr(data_delimiter)}'...")
                    final_cleaned_df = extracted_section_series.str.split(data_delimiter, expand=True)
                    num_header_cols = len(actual_headers)
                    num_data_cols = final_cleaned_df.shape[1]
                    
                    if num_data_cols > num_header_cols:
                        print(f"      -> AVISO: Mais colunas de dados ({num_data_cols}) que cabeçalhos ({num_header_cols}). Mantendo apenas as primeiras {num_header_cols}.")
                        final_cleaned_df = final_cleaned_df.iloc[:, :num_header_cols]
                    elif num_data_cols < num_header_cols:
                        print(f"      -> AVISO: Menos colunas de dados ({num_data_cols}) que cabeçalhos ({num_header_cols}). Adicionando colunas vazias.")
                        for i in range(num_data_cols, num_header_cols):
                            final_cleaned_df[i] = pd.NA
                    final_cleaned_df.columns = actual_headers
                    print(f"      -> Divisão inicial OK. DataFrame com {final_cleaned_df.shape[0]} linhas e {final_cleaned_df.shape[1]} colunas.")

                    original_file_headers = list(actual_headers) 
                    print(f"      -> Desmembrando data/hora (separador: '{date_time_separator}')...")
                    new_column_data = {}
                    cols_actually_dropped = []
                    
                    for original_dt_col_name, standardized_base_name in selected_lang_config["datetime_cols_map"].items():
                        if original_dt_col_name in final_cleaned_df.columns:
                            print(f"         -> Processando '{original_dt_col_name}' para base '{standardized_base_name}'")
                            col_as_str = final_cleaned_df[original_dt_col_name].astype(str).fillna('')
                            split_data = col_as_str.str.split(date_time_separator, n=1, expand=True)
                            if split_data.shape[1] == 1:
                                split_data[1] = pd.NA
                            date_col_name = f"Data {standardized_base_name}"
                            time_col_name = f"Hora {standardized_base_name}"

                            new_column_data[date_col_name] = split_data[0].replace({'': pd.NA, 'nan': pd.NA, 'NaT': pd.NA})
                            new_column_data[time_col_name] = split_data[1].replace({'': pd.NA, 'nan': pd.NA, 'NaT': pd.NA})
                            print(f"           -> Criadas '{date_col_name}', '{time_col_name}'")
                            cols_actually_dropped.append(original_dt_col_name)
                        else:
                            print(f"         -> AVISO: Coluna de data/hora original '{original_dt_col_name}' não encontrada no DataFrame para desmembrar.")
                    
                    for new_col, data in new_column_data.items():
                        final_cleaned_df[new_col] = data
                    
                    if cols_actually_dropped:
                        print(f"      -> Removendo colunas originais: {cols_actually_dropped}")
                        final_cleaned_df = final_cleaned_df.drop(columns=cols_actually_dropped)

                    final_cleaned_df.rename(columns=selected_lang_config["rename_to_pt_map"], inplace=True)
                    print(f"      -> Colunas renomeadas para padrão PT (se aplicável). Colunas atuais: {list(final_cleaned_df.columns)}")

                    final_column_order = []
                    processed_original_cols = set(cols_actually_dropped)
                    added_new_cols = set(new_column_data.keys())
                    for original_header in original_file_headers:
                        if original_header not in processed_original_cols:
                            renamed_header = selected_lang_config["rename_to_pt_map"].get(original_header, original_header)
                            if renamed_header in final_cleaned_df.columns:
                                final_column_order.append(renamed_header)
                        else: 
                            standardized_base = selected_lang_config["datetime_cols_map"][original_header]
                            date_col = f"Data {standardized_base}"
                            time_col = f"Hora {standardized_base}"
                            if date_col in added_new_cols: final_column_order.append(date_col)
                            if time_col in added_new_cols: final_column_order.append(time_col)
                    
                    for new_col in added_new_cols:
                        if new_col not in final_column_order:
                            final_column_order.append(new_col)
                    
                    final_cleaned_df = final_cleaned_df.reindex(columns=final_column_order)
                    print("      -> Desmembramento e reordenação OK.")
                    print("      -> Limpeza final (removendo linhas totalmente vazias)...")
                    final_cleaned_df = final_cleaned_df.dropna(how='all').reset_index(drop=True)
                    print(f"      -> Limpeza concluída: {len(final_cleaned_df)} linhas restantes.")
                
                elif extracted_section_series.empty:
                    print("      -> AVISO: Nenhuma linha de dados encontrada entre os marcadores.")
                    final_cleaned_df = pd.DataFrame()
                else:
                    raise ValueError("Não foi possível extrair cabeçalhos válidos da linha esperada.")
            else:
                raise ValueError(f"Índices de extração inválidos ou fora dos limites.\nInício: {start_row_index}, Fim: {end_row_index}, Cabeçalho: {header_row_index}, Dados: {data_start_index}, Total Linhas: {len(df_full)}")
            
            if final_cleaned_df is not None and not final_cleaned_df.empty:
                print(f"\n   > Salvando dados limpos em: {output_xlsx_path}")
                final_cleaned_df.to_excel(output_xlsx_path, sheet_name='Presenca Limpa', index=False, engine='openpyxl')
                print("   > Arquivo XLSX salvo com sucesso!")
                msg_success = f"Arquivo CSV limpo e salvo como:\n{output_xlsx_path}\n\nUse este novo arquivo XLSX no botão 'Anexar Lista(s) de Presença'."
                if messagebox.askyesno("Sucesso", msg_success + "\n\nDeseja abrir a pasta onde o arquivo foi salvo?", parent=self.root):
                    try:
                        subprocess.run(['explorer', os.path.normpath(output_folder)], check=False)
                    except FileNotFoundError:
                        try:
                            subprocess.run(['xdg-open', output_folder], check=False)
                        except FileNotFoundError:
                            try:
                                subprocess.run(['open', output_folder], check=False)
                            except FileNotFoundError:
                                print("   <!> Não foi possível abrir a pasta automaticamente.")
                    except Exception as e_open:
                        print(f"   <!> Erro ao tentar abrir a pasta: {e_open}")
            elif final_cleaned_df is not None and final_cleaned_df.empty:
                messagebox.showwarning("Aviso", "Nenhuma linha de dados válida encontrada na seção do CSV delimitada pelos marcadores.\nNenhum arquivo XLSX foi gerado.", parent=self.root)
            else:
                messagebox.showerror("Erro", "Ocorreu um erro inesperado e o DataFrame final não foi criado.", parent=self.root)
        
        except ValueError as ve:
            error_msg = f"Erro durante a limpeza do CSV:\n{ve}"
            print(f"   <!> ERRO: {error_msg}")
            messagebox.showerror("Erro Limpeza CSV", error_msg, parent=self.root)
        except KeyError as ke:
            error_msg = f"Erro: Coluna essencial não encontrada durante o processamento:\n{ke}"
            print(f"   <!> ERRO: {error_msg}")
            messagebox.showerror("Erro Coluna CSV", error_msg, parent=self.root)
        except Exception as e:
            error_details = traceback.format_exc()
            error_msg = f"Ocorreu um erro inesperado durante a limpeza do CSV:\n{type(e).__name__}: {e}"
            print(f"   <!> ERRO INESPERADO:\n{error_details}")
            messagebox.showerror("Erro Inesperado", error_msg + "\n\nVerifique o console para mais detalhes.", parent=self.root)
        print("--- Fim Limpeza CSV ---")

    def preparar_accounting_journey(self):
        print("\n--- Iniciando Preparação de Planilha (Accounting Journey) ---")
        input_header_row = 4
        col_nome_in = 'NOME'
        col_email_in = 'EMAIL/LOGIN'
        col_entrada_in = 'PRIMEIRO SINAL'
        col_saida_in = 'ÚLTIMO SINAL'
        col_nome_out = 'Nome'
        col_email_out = 'Email'
        col_data_ent_out = 'Data Entrada'
        col_hora_ent_out = 'Hora Entrada'
        col_data_sai_out = 'Data Saída'
        col_hora_sai_out = 'Hora Saída'
        col_duracao_out = 'Duração'
        col_funcao_out = 'Função'
        output_columns_order = [col_nome_out, col_data_ent_out, col_hora_ent_out, col_data_sai_out, col_hora_sai_out, col_duracao_out, col_email_out, col_funcao_out]
        datetime_format_in = '%d/%m/%Y %H:%M:%S'
        date_format_out = '%m/%d/%y'
        time_format_out = '%I:%M:%S %p'
        input_excel_path = filedialog.askopenfilename(
            title="Selecione a Planilha Bruta (Accounting Journey)",
            filetypes=[("Arquivos Excel", "*.xlsx;*.xls"), ("Todos os arquivos", "*.*")],
            parent=self.root
        )
        if not input_excel_path:
            print("   > Nenhuma arquivo selecionado.")
            messagebox.showinfo("Cancelado", "Nenhuma planilha selecionada.", parent=self.root)
            return
        if not input_excel_path.lower().endswith(('.xlsx', '.xls')):
            messagebox.showerror("Erro", "Arquivo inválido. Por favor, selecione um arquivo Excel (.xlsx ou .xls).", parent=self.root)
            return
        input_filename = os.path.basename(input_excel_path)
        output_folder = os.path.dirname(input_excel_path)
        base_filename_out = os.path.splitext(input_filename)[0]
        output_xlsx_path = os.path.join(output_folder, f"{base_filename_out}_preparada_para_importar.xlsx")
        print(f"   > Arquivo selecionado: {input_filename}")
        print(f"   > Saída formatada será: {os.path.basename(output_xlsx_path)}")
        try:
            print(f"   > Lendo planilha Excel (cabeçalho na linha {input_header_row + 1})...")
            try:
                df_input = pd.read_excel(input_excel_path, header=input_header_row)
                df_input.dropna(axis=1, how='all', inplace=True)
                df_input.dropna(axis=0, how='all', inplace=True)
                print(f"   > Leitura OK: {len(df_input)} linhas de dados encontradas.")
                print(f"   > Colunas encontradas: {list(df_input.columns)}")
            except FileNotFoundError:
                raise ValueError(f"Arquivo não encontrado: {input_filename}")
            except Exception as e_read:
                raise ValueError(f"Erro ao ler o arquivo Excel '{input_filename}': {e_read}")
            print("   > Verificando colunas necessárias...")
            input_cols_map = {}
            required_cols_in = {
                'nome': col_nome_in,
                'email': col_email_in,
                'entrada': col_entrada_in,
                'saida': col_saida_in
            }
            missing = []
            for key, expected_name in required_cols_in.items():
                found_col = find_col_ignore_case(df_input.columns, [expected_name])
                if found_col:
                    input_cols_map[key] = found_col
                    print(f"      -> Coluna '{key}' encontrada como: '{found_col}'")
                else:
                    missing.append(expected_name)
            if missing:
                raise ValueError(f"Colunas essenciais não encontradas na planilha de entrada: {', '.join(missing)}. Verifique o cabeçalho na linha {input_header_row + 1}.")
            print("   > Processando dados (extraindo Nome, Email, convertendo Data/Hora)...")
            df_output = pd.DataFrame()
            df_output[col_nome_out] = df_input[input_cols_map['nome']].astype(str).str.strip()
            df_output[col_email_out] = df_input[input_cols_map['email']].astype(str).str.lower().str.strip()
            print(f"   > Convertendo '{input_cols_map['entrada']}' para datetime (formato esperado: {datetime_format_in})...")
            dt_entrada = pd.to_datetime(df_input[input_cols_map['entrada']], format=datetime_format_in, errors='coerce')
            print(f"   > Convertendo '{input_cols_map['saida']}' para datetime (formato esperado: {datetime_format_in})...")
            dt_saida = pd.to_datetime(df_input[input_cols_map['saida']], format=datetime_format_in, errors='coerce')
            entrada_errors = dt_entrada.isna().sum()
            saida_errors = dt_saida.isna().sum()
            if entrada_errors > 0 or saida_errors > 0:
                print(f"   <!> AVISO: {entrada_errors} erros ao converter data/hora de entrada.")
                print(f"   <!> AVISO: {saida_errors} erros ao converter data/hora de saída.")
                print(f"       Verifique se o formato '{datetime_format_in}' corresponde aos dados nessas linhas.")
            print(f"   > Formatando Data Entrada ({date_format_out}) e Hora Entrada ({time_format_out})...")
            df_output[col_data_ent_out] = dt_entrada.dt.strftime(date_format_out)
            df_output[col_hora_ent_out] = dt_entrada.dt.strftime(time_format_out)
            print(f"   > Formatando Data Saída ({date_format_out}) e Hora Saída ({time_format_out})...")
            df_output[col_data_sai_out] = dt_saida.dt.strftime(date_format_out)
            df_output[col_hora_sai_out] = dt_saida.dt.strftime(time_format_out)
            df_output[col_duracao_out] = ""
            df_output[col_funcao_out] = ""
            df_output = df_output[output_columns_order]
            df_output.dropna(subset=[col_email_out], inplace=True)
            df_output = df_output[df_output[col_email_out].str.contains('@', na=False)]
            initial_rows = len(df_output)
            df_output.dropna(subset=[col_data_ent_out, col_hora_ent_out, col_data_sai_out, col_hora_sai_out], how='any', inplace=True)
            final_rows = len(df_output)
            if initial_rows > final_rows:
                print(f"   > {initial_rows - final_rows} linhas removidas por falta de data/hora de entrada ou saída válidas.")
            if df_output.empty:
                raise ValueError("Nenhuma linha válida restou após o processamento. Verifique os dados de entrada, especialmente as datas/horas.")
            print(f"   > Processamento concluído. {len(df_output)} linhas válidas formatadas.")
            print(f"\n   > Salvando planilha formatada em: {output_xlsx_path}")
            df_output.to_excel(output_xlsx_path, sheet_name='Presenca Formatada', index=False, engine='openpyxl')
            print("   > Arquivo XLSX formatado salvo com sucesso!")
            msg_success = f"Planilha '{input_filename}' processada com sucesso!\n\nArquivo formatado salvo como:\n{output_xlsx_path}\n\nUse este novo arquivo no botão 'Anexar Lista(s) de Presença' (para modo Digital/Híbrido)."
            if messagebox.askyesno("Sucesso", msg_success + "\n\nDeseja abrir a pasta onde o arquivo foi salvo?", parent=self.root):
                try:
                    subprocess.run(['explorer', os.path.normpath(output_folder)], check=False)
                except FileNotFoundError:
                    try:
                        subprocess.run(['xdg-open', output_folder], check=False)
                    except FileNotFoundError:
                        try:
                            subprocess.run(['open', output_folder], check=False)
                        except FileNotFoundError:
                            print("   <!> Não foi possível abrir a pasta automaticamente.")
                except Exception as e_open:
                    print(f"   <!> Erro ao tentar abrir a pasta: {e_open}")
        except ValueError as ve:
            error_msg = f"Erro durante a preparação da planilha:\n{ve}"
            print(f"   <!> ERRO (ValueError): {error_msg}")
            messagebox.showerror("Erro na Preparação", error_msg, parent=self.root)
        except KeyError as ke:
            error_msg = f"Erro: Coluna não encontrada durante o processamento:\n{ke}\nVerifique os nomes das colunas no arquivo de entrada."
            print(f"   <!> ERRO (KeyError): {error_msg}")
            messagebox.showerror("Erro Coluna", error_msg, parent=self.root)
        except Exception as e:
            error_details = traceback.format_exc()
            error_msg = f"Ocorreu um erro inesperado durante a preparação:\n{type(e).__name__}: {e}"
            print(f"   <!> ERRO INESPERADO:\n{error_details}")
            messagebox.showerror("Erro Inesperado", error_msg + "\n\nVerifique o console para mais detalhes.", parent=self.root)
        finally:
            print("--- Fim Preparação Planilha (Accounting Journey) ---")

    def processar_nps(self, file_path):
        if not file_path or not os.path.exists(file_path):
            return "N/A", "Arquivo NPS não fornecido.", ""
        print(f"--- Processando NPS (v3 - Totais/Regex Flexível): {os.path.basename(file_path)} ---")
        texto_extraido = ""
        try:
            with pdfplumber.open(file_path) as pdf:
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text(x_tolerance=2, y_tolerance=3, layout=True)
                    if page_text:
                        texto_extraido += f"--- Página {i+1} ---\n{page_text}\n\n"
                    else:
                        page_text_simple = page.extract_text(x_tolerance=1, y_tolerance=3)
                        if page_text_simple:
                            texto_extraido += f"--- Página {i+1} (Simples) ---\n{page_text_simple}\n\n"
            if not texto_extraido.strip():
                print(" <!> AVISO: Nenhum texto extraído do PDF.")
                return "N/A", "Texto não extraído do PDF.", ""
            print("   > Texto extraído. Procurando totais NPS (Promoters, Passives, Detractors)...")
            promoters = None
            passives = None
            detractors = None
            flags = re.IGNORECASE | re.MULTILINE
            promoter_pattern = re.compile(r"Promoters(?:\s|\n)+(\d+)", flags)
            passive_pattern = re.compile(r"Passives(?:\s|\n)+(\d+)", flags)
            detractor_pattern= re.compile(r"Detractors(?:\s|\n)+(\d+)", flags)
            promoter_match = promoter_pattern.search(texto_extraido)
            passive_match  = passive_pattern.search(texto_extraido)
            detractor_match= detractor_pattern.search(texto_extraido)
            try:
                if promoter_match:
                    promoters = int(promoter_match.group(1))
                    print(f"      > Promoters encontrados: {promoters}")
                else:
                    print("      <!> Padrão 'Promoters... número' não encontrado no texto.")
                if passive_match:
                    passives = int(passive_match.group(1))
                    print(f"      > Passives encontrados: {passives}")
                else:
                    print("      <!> Padrão 'Passives... número' não encontrado no texto.")
                if detractor_match:
                    detractors = int(detractor_match.group(1))
                    print(f"      > Detractors encontrados: {detractors}")
                else:
                    print("      <!> Padrão 'Detractors... número' não encontrado no texto.")
            except ValueError as e_val:
                print(f"   <!> ERRO: Valor inválido encontrado para Promoters/Passives/Detractors: {e_val}")
                return "Erro Conversão", "Valor inválido nos totais NPS.", texto_extraido
            except Exception as e_re:
                print(f"   <!> ERRO durante busca regex NPS: {e_re}")
                return "Erro Regex", f"Erro regex NPS: {e_re}", texto_extraido
            nps_score_str = "N/A"
            nps_details_str = "Totais P/Pa/D não encontrados ou incompletos."
            if promoters is not None and passives is not None and detractors is not None:
                total_respondents = promoters + passives + detractors
                nps_details_str = f"Promoters: {promoters}, Passives: {passives}, Detractors: {detractors} (Total: {total_respondents})"
                if total_respondents > 0:
                    nps_score = ((promoters - detractors) / total_respondents) * 100
                    nps_score_str = f"{nps_score:.2f}%"
                    print(f"   > NPS Calculado: Score={nps_score_str}, {nps_details_str}")
                else:
                    nps_score_str = "0.00%"
                    print(f"   > Total de respondentes é 0. NPS considerado {nps_score_str}.")
            else:
                print(f"   > Não foi possível calcular o NPS pois um ou mais totais (Promoters, Passives, Detractors) não foram encontrados.")
            return nps_score_str, nps_details_str, texto_extraido
        except Exception as e:
            print(f"Erro CRÍTICO ao processar PDF NPS '{os.path.basename(file_path)}':\n{traceback.format_exc()}")
            return "Erro Leitura PDF", f"Erro crítico ao ler PDF: {e}", ""

    def verificar_divergencias(self, convidados_df, presenca_df):
        convidados_emails = set()
        presenca_emails = set()
        
        email_col_conv = find_col_ignore_case(convidados_df.columns, ['email', 'e-mail', 'correio eletrônico'])
        if email_col_conv:
            convidados_emails = set(convidados_df[email_col_conv].dropna().astype(str).str.lower().str.strip())
            convidados_emails.discard('')
            print(f"   > Divergências: {len(convidados_emails)} e-mails de convidados encontrados para verificação.")
        else:
            print(f"   <!> AVISO Divergências: Coluna de e-mail não encontrada nos convidados.")

        email_col_pres = find_col_ignore_case(presenca_df.columns, ['email', 'e-mail'])
        if email_col_pres:
            presenca_emails = set(presenca_df[email_col_pres].dropna().astype(str).str.lower().str.strip())
            presenca_emails.discard('')
            print(f"   > Divergências: {len(presenca_emails)} e-mails de presença encontrados para verificação.")
        else:
            print(f"   <!> AVISO Divergências: Coluna de e-mail não encontrada na presença.")

        if not convidados_emails and not presenca_emails:
            print("   > Divergências: Impossível verificar, e-mails não encontrados em nenhuma das listas.")
            return [], []

        nao_presentes = sorted(list(convidados_emails - presenca_emails))
        extras_na_presenca = sorted(list(presenca_emails - convidados_emails))
        
        if nao_presentes:
            print(f"     - {len(nao_presentes)} Convidado(s) que não compareceu(ram) encontrado(s).")
        if extras_na_presenca:
            print(f"     - {len(extras_na_presenca)} Presente(s) não convidado(s) encontrado(s).")
            
        return nao_presentes, extras_na_presenca

    # <<<<<<< NOVA FUNCIONALIDADE: A função agora aceita um flag para mudar seu comportamento >>>>>>>
    def _criar_aba_divergencias(self, wb, nao_presentes, extras_na_presenca, is_master_list_mode=False):
        if not extras_na_presenca and (is_master_list_mode or not nao_presentes):
            print("   > Nenhuma divergência relevante encontrada para gerar a aba.")
            return

        print("   > Criando aba de 'Divergências' no relatório...")
        ws_div = wb.create_sheet(title="Divergências")
        
        header_font = Font(bold=True, size=12, color="FFFFFF")
        nao_presente_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
        extra_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
        
        col_idx_extra = 1
        # Se não estivermos no modo master list, a coluna de faltantes é criada
        if not is_master_list_mode:
            cell_a1 = ws_div.cell(row=1, column=1, value=f"Convidados que Faltaram ({len(nao_presentes)})")
            cell_a1.font = header_font
            cell_a1.fill = nao_presente_fill
            ws_div.column_dimensions['A'].width = 45
            for i, email in enumerate(nao_presentes, start=2):
                ws_div.cell(row=i, column=1, value=email)
            col_idx_extra = 2

        # A coluna de extras (presentes não convidados / não encontrados na base) é sempre criada
        header_text_extra = f"Presentes não encontrados na Base de Profissionais ({len(extras_na_presenca)})" if is_master_list_mode else f"Presentes não Convidados ({len(extras_na_presenca)})"
        
        cell_extra = ws_div.cell(row=1, column=col_idx_extra, value=header_text_extra)
        cell_extra.font = header_font
        cell_extra.fill = extra_fill
        
        col_letter_extra = 'B' if col_idx_extra == 2 else 'A'
        ws_div.column_dimensions[col_letter_extra].width = 45

        for i, email in enumerate(extras_na_presenca, start=2):
            ws_div.cell(row=i, column=col_idx_extra, value=email)
    # --- FIM DA NOVA FUNCIONALIDADE ---


    def gerar_relatorio(self):
        # <<<<<<< NOVA FUNCIONALIDADE: Validação da fonte de convidados >>>>>>>
        if self.use_professionals_report_var.get():
             if not self.master_professionals_path or not os.path.exists(self.master_professionals_path):
                messagebox.showerror("Erro", "O modo 'Relatório de Profissionais' está ativo, mas o caminho do arquivo não foi configurado ou é inválido.\n\nVá para a aba 'Configuração' para defini-lo.", parent=self.root)
                return
        else:
            if not self.convidados_path.get():
                messagebox.showerror("Erro", "Anexe a Lista de Convidados ou selecione a opção para usar o 'Relatório de Profissionais'.", parent=self.root)
                return
        # --- FIM DA NOVA FUNCIONALIDADE ---

        if not self.presenca_paths:
            messagebox.showerror("Erro", "Anexe pelo menos uma Lista de Presença.", parent=self.root)
            return
            
        try:
            self._get_instrutores_data() 
        except ValueError as e:
            messagebox.showerror("Erro de Validação (Instrutores)", str(e), parent=self.root)
            return

        tipo_curso = self.tipo_curso_var.get()
        if tipo_curso in ["Digital", "Híbrido"]:
            ch_entry_value = self.entries["Carga Horária Bruta Total (h)"].get().strip()
            if not ch_entry_value:
                messagebox.showerror("Erro Validação", f"Para tipo de curso '{tipo_curso}', o campo 'Carga Horária Bruta Total (h)' deve ser preenchido.\nUse o formato HH:MM (ex: 16:00) ou um número (ex: 16).", parent=self.root)
                return
        self.gerar_button.configure(state=tk.DISABLED)
        self.progress.set(0)
        self.root.update_idletasks()
        self.progress.start()
        self.root.after(150, self._gerar_relatorio_async)

    def _gerar_relatorio_async(self):
        try:
            tipo_curso = self.tipo_curso_var.get()
            print(f"\n--- Iniciando Geração de Relatório (Modo: {tipo_curso}) ---")
            if tipo_curso == "Digital":
                self.gerar_relatorio_digital()
            elif tipo_curso == "Híbrido":
                self.gerar_relatorio_hibrido()
            elif tipo_curso == "Presencial":
                self.gerar_relatorio_presencial()
            else:
                messagebox.showerror("Erro Interno", f"Tipo de curso desconhecido: {tipo_curso}", parent=self.root)
                print(f"ERRO: Tipo de curso inválido selecionado: {tipo_curso}")
        except Exception as e:
            error_details = traceback.format_exc()
            messagebox.showerror("Erro Fatal na Geração", f"Ocorreu um erro inesperado durante a geração do relatório:\n{type(e).__name__}: {e}\n\nVerifique os arquivos de entrada e as configurações.\nDetalhes no console.", parent=self.root)
            print(f"ERRO FATAL GERAÇÃO ({self.tipo_curso_var.get()}):\n{error_details}")
        finally:
            self.progress.stop()
            self.progress.set(0)
            self.root.update_idletasks()
            self.gerar_button.configure(state=tk.NORMAL)
            print(f"--- Finalizando Tentativa de Geração (Modo: {self.tipo_curso_var.get()}) ---")

    # <<<<<<< NOVA FUNCIONALIDADE: Nova função para ler a base de profissionais >>>>>>>
    def _read_professionals_report(self):
        file_path = self.master_professionals_path
        print(f"Lendo Relatório de Profissionais: {os.path.basename(file_path)}")
        
        try:
            df_initial = pd.read_excel(file_path, header=None)
        except Exception as e:
            raise ValueError(f"Erro ao tentar ler o Relatório de Profissionais '{os.path.basename(file_path)}': {e}")
        
        header_row = -1
        # Palavras-chave flexíveis para encontrar o cabeçalho
        keywords_header = ['nome', 'e-mail', 'cargo', 'email', 'name']
        for i, row in df_initial.head(20).iterrows():
            try:
                row_str = ' '.join(map(str, row.dropna().tolist())).lower()
                matches = [keyword for keyword in keywords_header if keyword in row_str]
                # Procuramos por pelo menos 'nome' e 'email'
                if len(matches) >= 2 and ('nome' in row_str or 'name' in row_str) and ('e-mail' in row_str or 'email' in row_str):
                    print(f"   > Possível header de profissionais encontrado na linha {i+1} (índice {i}).")
                    header_row = i
                    break
            except Exception:
                continue
        
        if header_row == -1:
            raise ValueError("Não foi possível encontrar a tabela de dados no Relatório de Profissionais. Verifique se colunas como 'Nome' e 'E-mail' existem.")

        df = pd.read_excel(file_path, header=header_row)
        df.columns = df.columns.astype(str).str.strip()
        print(f"   > Relatório de profissionais lido com sucesso. Colunas: {list(df.columns)}")
        
        # Adiciona uma coluna 'Inscrição' vazia para compatibilidade de merge, se não existir
        if 'Inscrição' not in df.columns:
            df['Inscrição'] = 'N/A (Base Geral)'

        return df
    # --- FIM DA NOVA FUNCIONALIDADE ---

    # <<<<<<< NOVA FUNCIONALIDADE: Nova função que centraliza a obtenção da base de participantes >>>>>>>
    def _get_base_participants_df(self):
        if self.use_professionals_report_var.get():
            print("Lendo dados do 'Relatório de Profissionais'...")
            return self._read_professionals_report()
        else:
            print(f"Lendo arquivo de Convidados anexado...")
            return self._read_convidados_file(self.convidados_path.get())
    # --- FIM DA NOVA FUNCIONALIDADE ---

    def _read_convidados_file(self, file_path):
        print(f"Lendo arquivo de Convidados: {os.path.basename(file_path)}")
        
        try:
            df_initial = pd.read_excel(file_path, header=None)
        except Exception as e:
            raise ValueError(f"Erro ao tentar ler o arquivo Excel de convidados '{os.path.basename(file_path)}': {e}")
        
        header_row = -1
        keywords_header = ['nome', 'cargo', 'e-mail', 'convocação', 'convite', 'bu']
        for i, row in df_initial.head(20).iterrows():
            try:
                row_str = ' '.join(map(str, row.dropna().tolist())).lower()
                matches = [keyword for keyword in keywords_header if keyword in row_str]
                if len(matches) >= 3:
                    print(f"   > Possível header de convidados encontrado na linha do Excel {i+1} (índice {i}).")
                    header_row = i
                    break
            except Exception:
                continue
        
        if header_row == -1:
            raise ValueError("Não foi possível encontrar a tabela de dados na Lista de Convidados. Verifique se colunas como 'NOME', 'CARGO' e 'E-MAIL' existem.")

        df = pd.read_excel(file_path, header=header_row)
        df.columns = df.columns.astype(str).str.strip()
        print(f"   > Tabela de convidados lida com sucesso. Colunas: {list(df.columns)}")
        
        convocacao_col = find_col_ignore_case(df.columns, ['convocação'])
        convite_col = find_col_ignore_case(df.columns, ['convite'])
        
        if convocacao_col and convite_col:
            print("   > Processando colunas 'Convocação' e 'Convite'.")
            df[convocacao_col] = df[convocacao_col].astype(str).str.strip().str.lower()
            df[convite_col] = df[convite_col].astype(str).str.strip().str.lower()
            df['Inscrição'] = 'Não especificado'
            df.loc[df[convocacao_col] == 'x', 'Inscrição'] = 'Convocado'
            df.loc[df[convite_col] == 'x', 'Inscrição'] = 'Convidado'
        else:
            print("   > AVISO: Colunas 'Convocação' ou 'Convite' não encontradas. A coluna 'Inscrição' não será criada.")
            df['Inscrição'] = 'N/A' 

        return df

    def _read_input_file(self, file_path):
        print(f"Lendo arquivo de Presença/Genérico: {os.path.basename(file_path)}")
        file_path_lower = file_path.lower()
        if file_path_lower.endswith(('.xlsx', '.xls')):
            try:
                df = pd.read_excel(file_path, header=0)
                df.columns = df.columns.astype(str).str.strip()
                print(f"   > Colunas lidas do Excel: {list(df.columns)}")
                return df
            except Exception as e:
                raise ValueError(f"Erro ao ler o arquivo Excel '{os.path.basename(file_path)}': {e}") from e
        elif file_path_lower.endswith('.csv'):
            try:
                df = read_csv_with_encoding(file_path)
                df.columns = df.columns.astype(str).str.strip()
                return df
            except ValueError as e_csv:
                raise ValueError(f"Falha ao tentar ler o CSV '{os.path.basename(file_path)}': {e_csv}") from e_csv
        else:
            raise ValueError(f"Formato de arquivo não suportado: {os.path.basename(file_path)}. Use XLSX, XLS ou CSV.")

    def _aplicar_estilos_excel(self, ws):
        header_font=Font(bold=True, color="FFFFFF", name='Calibri', size=11)
        header_fill=PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border_thin=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_alignment=Alignment(horizontal="right", vertical="center", wrap_text=False)
        section_header_font=Font(bold=True, size=12, color="4F81BD", name='Calibri')
        green_fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_font=Font(color="006100", name='Calibri', size=11)
        red_font=Font(color="9C0006", name='Calibri', size=11)
        italic_font=Font(italic=True, name='Calibri', size=10, color="808080")
        
        ws.column_dimensions['A'].width = 35 # Nome
        ws.column_dimensions['B'].width = 35 # Email
        ws.column_dimensions['C'].width = 18 # CPF
        ws.column_dimensions['D'].width = 20 # Contagem Presenca / Duração
        ws.column_dimensions['E'].width = 15 # % Presença
        ws.column_dimensions['F'].width = 15 # Status
        ws.column_dimensions['G'].width = 15 # Inscrição
        ws.column_dimensions['H'].width = 25 # Cargo
        ws.column_dimensions['I'].width = 20 # BU
        ws.column_dimensions['J'].width = 15 # Observação


        return (header_font, header_fill, border_thin, center_alignment, left_alignment, right_alignment, section_header_font, green_fill, red_fill, green_font, red_font, italic_font)

    def _processar_arquivo_presenca(self, file_path, id_col_primaria='Email'):
        base_filename = os.path.basename(file_path)
        print(f"      Processando arquivo individual (Híbrido/Aux): {base_filename}")
        e_digital = None
        try:
            if not file_path.lower().endswith(('.xlsx', '.xls')):
                print(f"         > '{base_filename}' não é Excel. Tentando lógica Presencial/CSV...")
                try:
                    presenca_df = self._read_input_file(file_path)
                    
                    id_col_pres = None
                    colunas_prioritarias_email = ["E-mail KPMG", "E-mail Corporativo", "Email", "e-mail", "user email"]
                    colunas_prioritarias_nome = ["Nome Completo", "Name", "Nome"]

                    opcoes_busca = colunas_prioritarias_email + colunas_prioritarias_nome
                    if id_col_primaria == 'Nome':
                        opcoes_busca = colunas_prioritarias_nome + colunas_prioritarias_email

                    for col_opt in opcoes_busca:
                        id_col_pres = find_col_ignore_case(presenca_df.columns, [col_opt])
                        if id_col_pres:
                            print(f"           [CSV] Encontrada coluna de ID prioritária: '{id_col_pres}' (Opção: '{col_opt}')")
                            break
                    
                    if not id_col_pres:
                        raise ValueError(f"Nenhuma coluna de ID utilizável (tentativas: {', '.join(opcoes_busca)}) encontrada em '{base_filename}'.")

                    ids_presentes = set(presenca_df[id_col_pres].dropna().astype(str).str.strip())
                    
                    is_email_column = any(keyword in id_col_pres.lower() for keyword in colunas_prioritarias_email)
                    if is_email_column:
                        ids_presentes = {s.lower() for s in ids_presentes if s and '@' in s}
                    else:
                        ids_presentes.discard('')

                    if not ids_presentes:
                        raise ValueError(f"Nenhum ID válido extraído da coluna '{id_col_pres}' no arquivo CSV '{base_filename}'.")
                    
                    return 'presencial', ids_presentes
                except Exception as e_pres_fallback:
                    print(f"         <!> Falha PRESENCIAL (fallback): {e_pres_fallback}.")
                    return 'falha', f"{base_filename} (Não Excel e falha Presencial: {e_pres_fallback})"
            
            presenca_df = self._read_input_file(file_path)
            cols_digital = {"email": ['email', 'e-mail'], "data_entrada": ['data entrada'], "hora_entrada": ['hora entrada'], "data_saida": ['data saída'], "hora_saida": ['hora saída']}
            found_cols_digital = {key: find_col_ignore_case(presenca_df.columns, names) for key, names in cols_digital.items()}
            
            if all(found_cols_digital.values()):
                print(f"         > Colunas digitais encontradas. Tentando lógica DIGITAL...")
                try:
                    presenca_df_copy = presenca_df.copy()
                    presenca_df_copy.rename(columns={v: k for k, v in found_cols_digital.items()}, inplace=True)
                    for col in ['data_entrada', 'hora_entrada', 'data_saida', 'hora_saida', 'email']:
                        presenca_df_copy[col] = presenca_df_copy[col].astype(str).str.strip()
                    presenca_df_copy.dropna(subset=['email'], inplace=True)
                    presenca_df_copy = presenca_df_copy[presenca_df_copy['email'].str.contains('@', na=False)]
                    if presenca_df_copy.empty:
                        raise ValueError("Nenhum email válido encontrado após limpeza (digital).")
                    
                    dt_formats = ['%m/%d/%y %H:%M', '%d/%m/%y %H:%M', '%m/%d/%Y %H:%M', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M', '%m/%d/%y %I:%M %p', '%d/%m/%y %I:%M %p', '%m/%d/%Y %I:%M %p', '%d/%m/%Y %I:%M %p', '%m/%d/%y %I:%M:%S %p', '%d/%m/%y %I:%M:%S %p', '%m/%d/%Y %I:%M:%S %p', '%d/%m/%Y %I:%M:%S %p', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%d/%m/%y %H:%M:%S', '%m/%d/%y %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f']
                    
                    def parse_dt(date_str, time_str, formats):
                        d_str = str(date_str).split(' ')[0].strip() if pd.notna(date_str) else ""
                        t_str = str(time_str).strip() if pd.notna(time_str) else ""
                        if not d_str or d_str.lower() == 'nan' or not t_str or t_str.lower() == 'nan': return pd.NaT
                        combined_str = f"{d_str} {t_str}"
                        for fmt in formats:
                            try: return pd.to_datetime(combined_str, format=fmt)
                            except (ValueError, TypeError): continue
                        try: return pd.to_datetime(combined_str, errors='coerce')
                        except (ValueError, TypeError): return pd.NaT

                    presenca_df_copy['Entrada_dt']=presenca_df_copy.apply(lambda r: parse_dt(r['data_entrada'], r['hora_entrada'], dt_formats), axis=1)
                    presenca_df_copy['Saida_dt']=presenca_df_copy.apply(lambda r: parse_dt(r['data_saida'], r['hora_saida'], dt_formats), axis=1)
                    presenca_df_copy.dropna(subset=['Entrada_dt', 'Saida_dt'], inplace=True)
                    presenca_df_copy = presenca_df_copy[presenca_df_copy['Saida_dt'] > presenca_df_copy['Entrada_dt']].copy()
                    if presenca_df_copy.empty:
                        raise ValueError(f"Nenhuma linha com datas/horas válidas.")
                    
                    presenca_df_copy['Duracao_horas']=(presenca_df_copy['Saida_dt']-presenca_df_copy['Entrada_dt']).dt.total_seconds()/3600.0
                    digital_df = presenca_df_copy[['email', 'Duracao_horas']].rename(columns={'email':'Email'}).copy()
                    return 'digital', digital_df
                except Exception as e_dig_inner:
                    e_digital = e_dig_inner
                    print(f"         <!> Falha na lógica DIGITAL: {e_digital}. Tentando Presencial como fallback...")
            
            print(f"         > Tentando lógica PRESENCIAL (fallback para Excel)...")
            try:
                id_col_pres = None
                colunas_prioritarias_email = ["E-mail KPMG", "E-mail Corporativo", "Email", "e-mail"]
                colunas_prioritarias_nome = ["Nome Completo", "Name", "Nome"]

                opcoes_busca = colunas_prioritarias_email + colunas_prioritarias_nome
                if id_col_primaria == 'Nome':
                    opcoes_busca = colunas_prioritarias_nome + colunas_prioritarias_email

                for col_opt in opcoes_busca:
                    id_col_pres = find_col_ignore_case(presenca_df.columns, [col_opt])
                    if id_col_pres:
                        print(f"           [Excel Fallback] Encontrada coluna de ID prioritária: '{id_col_pres}'")
                        break
                
                if not id_col_pres:
                    raise ValueError(f"Nenhuma coluna de ID utilizável (tentativas: {', '.join(opcoes_busca)}) encontrada em '{base_filename}'.")

                ids_presentes = set(presenca_df[id_col_pres].dropna().astype(str).str.strip())
                
                is_email_column = any(keyword in id_col_pres.lower() for keyword in colunas_prioritarias_email)
                if is_email_column:
                    ids_presentes = {s.lower() for s in ids_presentes if s and '@' in s}
                else:
                    ids_presentes.discard('')

                if not ids_presentes:
                    raise ValueError(f"Nenhum ID válido extraído da coluna '{id_col_pres}' (fallback presencial Excel).")
                
                return 'presencial', ids_presentes
            except Exception as e_presencial_excel:
                erro_digital_msg = f"Erro Digital: {e_digital}" if e_digital else "Colunas Digitais Ausentes"
                return 'falha', f"{base_filename} ({erro_digital_msg} | Erro Presencial Fallback Excel: {e_presencial_excel})"
        except Exception as e_geral:
            return 'falha', f"{base_filename} (Erro geral inesperado: {e_geral})"

    def gerar_relatorio_digital(self):
        try:
            print("\n--- Iniciando Geração Relatório Digital ---")
            self.progress.set(0.05)
            self.root.update_idletasks()
            
            # <<<<<<< NOVA FUNCIONALIDADE: Usa a nova função para obter a base de participantes >>>>>>>
            print("1. Obtendo base de participantes...")
            convidados_df = self._get_base_participants_df()
            # --- FIM DA NOVA FUNCIONALIDADE ---
            
            email_col_conv = find_col_ignore_case(convidados_df.columns, ['email', 'e-mail'])
            if not email_col_conv:
                raise ValueError("Coluna de Email não encontrada na lista de participantes (convidados ou profissionais).")
            if email_col_conv != 'Email':
                 convidados_df.rename(columns={email_col_conv: 'Email'}, inplace=True)
            
            convidados_df.dropna(subset=['Email'], inplace=True)
            convidados_df['Email'] = convidados_df['Email'].astype(str).str.lower().str.strip()
            convidados_df = convidados_df[convidados_df['Email'].str.contains('@', na=False)]
            if convidados_df.empty:
                raise ValueError("Nenhum email válido na lista de participantes.")
            
            print("   > Obtendo dados dos instrutores...")
            instrutores_data, instrutores_emails = self._get_instrutores_data()
            instrutores_emails_set = set(instrutores_emails)
            if instrutores_emails_set:
                print(f"     - Instrutores a serem sinalizados: {instrutores_emails_set}")

            self.progress.set(0.15)
            self.root.update_idletasks()
            
            print("\n2. Processando Arquivos de Presença (Modo Digital)...")
            all_presenca_entries = []
            arquivos_falha = []
            col_names_needed = {
                "email": ['email', 'e-mail', 'correio eletrônico', 'user email'],
                "data_entrada": ['data entrada', 'join date'], "hora_entrada": ['hora entrada', 'join time'],
                "data_saida": ['data saída', 'leave date'], "hora_saida": ['hora saída', 'leave time']
            }
            datetime_formats_to_try = [
                '%m/%d/%y %H:%M', '%d/%m/%y %H:%M', '%m/%d/%Y %H:%M', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M',
                '%m/%d/%y %I:%M %p', '%d/%m/%y %I:%M %p', '%m/%d/%Y %I:%M %p', '%d/%m/%Y %I:%M %p',
                '%m/%d/%y %I:%M:%S %p', '%d/%m/%y %I:%M:%S %p', '%m/%d/%Y %I:%M:%S %p', '%d/%m/%Y %I:%M:%S %p',
                '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%d/%m/%y %H:%M:%S', '%m/%d/%y %H:%M:%S',
                '%Y-%m-%d %H:%M:%S.%f'
            ]

            def parse_datetime_multiformat(date_str, time_str, formats):
                d_str = str(date_str).split(' ')[0].strip() if pd.notna(date_str) else ""
                t_str = str(time_str).strip() if pd.notna(time_str) else ""
                if not d_str or d_str.lower() == 'nan' or not t_str or t_str.lower() == 'nan': return pd.NaT
                combined_str = f"{d_str} {t_str}"
                for fmt in formats:
                    try: return pd.to_datetime(combined_str, format=fmt)
                    except (ValueError, TypeError): continue
                try: return pd.to_datetime(combined_str, errors='coerce')
                except (ValueError, TypeError): return pd.NaT
                    
            for file_path in self.presenca_paths:
                try:
                    presenca_df = self._read_input_file(file_path)
                    found_cols = {key: find_col_ignore_case(presenca_df.columns, names) for key, names in col_names_needed.items()}
                    if not all(found_cols.values()):
                        missing_desc = [f"'{'/'.join(col_names_needed[m])}'" for m, f in found_cols.items() if not f]
                        raise ValueError(f"Colunas essenciais não encontradas: {', '.join(missing_desc)}.")
                    
                    presenca_df.rename(columns={v: k for k, v in found_cols.items()}, inplace=True)
                    presenca_df['Entrada_dt'] = presenca_df.apply(lambda row: parse_datetime_multiformat(row['data_entrada'], row['hora_entrada'], datetime_formats_to_try), axis=1)
                    presenca_df['Saida_dt'] = presenca_df.apply(lambda row: parse_datetime_multiformat(row['data_saida'], row['hora_saida'], datetime_formats_to_try), axis=1)
                    presenca_df.dropna(subset=['Entrada_dt', 'Saida_dt', 'email'], inplace=True)
                    presenca_df['Duracao_horas'] = (presenca_df['Saida_dt'] - presenca_df['Entrada_dt']).dt.total_seconds() / 3600.0
                    
                    presenca_valid_df = presenca_df[presenca_df['Duracao_horas'] > 0][['email', 'Duracao_horas']].copy()
                    presenca_valid_df['email'] = presenca_valid_df['email'].astype(str).str.lower().str.strip()
                    all_presenca_entries.append(presenca_valid_df)
                except Exception as e:
                    arquivos_falha.append(f"{os.path.basename(file_path)} (Erro: {e})")

            if not all_presenca_entries:
                raise ValueError("Nenhum dado de presença válido pôde ser extraído dos arquivos.")

            presenca_total_df = pd.concat(all_presenca_entries, ignore_index=True)
            
            net_training_hours, total_gross_hours, total_interval_hours, daily_gross_hours, daily_interval_hours = self._get_effective_training_hours()
            
            presenca_agregada = presenca_total_df.groupby("email", as_index=False).agg(
                Duracao_Total_Horas=("Duracao_horas", "sum")
            ).rename(columns={'email': 'Email'})
            
            print(f"   > Deduzindo o tempo total de intervalo ({total_interval_hours:.2f}h) da duração de cada participante.")
            presenca_agregada["Duracao_Total_Horas"] = (presenca_agregada["Duracao_Total_Horas"] - total_interval_hours).clip(lower=0)
            
            presenca_agregada["% Presença"] = (presenca_agregada["Duracao_Total_Horas"] / net_training_hours * 100).clip(upper=100.0)

            print("\n3. Combinando dados e preparando o relatório final...")
            
            final_col_order = ['Nome', 'Email', 'CPF', 'Duracao_Total_Horas', '% Presença', 'Status', 'Inscrição', 'Cargo', 'BU', 'Observação']
            cols_to_merge = ['Email', 'Nome', 'Inscrição']
            
            map_convidados = {'CPF': ['cpf'], 'Cargo': ['cargo'], 'BU': ['bu'], 'Nome': ['nome']}
            for std_name, possible_names in map_convidados.items():
                found = find_col_ignore_case(convidados_df.columns, possible_names)
                if found and std_name not in convidados_df.columns:
                    convidados_df.rename(columns={found: std_name}, inplace=True)
                if std_name in convidados_df.columns and std_name not in cols_to_merge:
                    cols_to_merge.append(std_name)
            
            cols_cv_existentes = [c for c in cols_to_merge if c in convidados_df.columns]
            
            # <<<<<<< NOVA FUNCIONALIDADE: Lógica de merge condicional >>>>>>>
            # Se o checkbox estiver marcado, faz um 'inner' merge (só presentes). Caso contrário, 'left' (todos os convidados).
            merge_how = "inner" if self.use_professionals_report_var.get() else "left"
            print(f"   > Usando modo de merge: '{merge_how}'.")
            merged_df = pd.merge(convidados_df.drop_duplicates(subset=['Email'], keep='first')[cols_cv_existentes], presenca_agregada, on="Email", how=merge_how)
            # --- FIM DA NOVA FUNCIONALIDADE ---
            
            merged_df["Duracao_Total_Horas"].fillna(0, inplace=True)
            merged_df["% Presença"].fillna(0, inplace=True)
            merged_df["Status"] = merged_df["% Presença"].apply(lambda x: "Presente" if x >= 75 else "Falta")
            merged_df['Observação'] = ''
            if instrutores_emails_set:
                merged_df.loc[merged_df['Email'].isin(instrutores_emails_set), 'Observação'] = 'Instrutor'
            
            for col in final_col_order:
                if col not in merged_df.columns: merged_df[col] = ""

            merged_df = merged_df[final_col_order].sort_values(by="Nome", key=lambda col: col.astype(str).str.lower())
            
            self.progress.set(0.6)
            self.root.update_idletasks()
            
            print("\n4. Verificando Divergências...")
            nao_presentes, extras_na_presenca = self.verificar_divergencias(convidados_df, presenca_agregada)

            print("\n5. Processando NPS...")
            nps_path = self.nps_path.get()
            nps_score_str, nps_details_str, nps_texto_completo = self.processar_nps(nps_path) if nps_path and os.path.exists(nps_path) else ("N/A", "Arquivo não fornecido.", "")
            
            print("\n6. Gerando Arquivo Excel (Digital)...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório Presença"
            styles_excel = self._aplicar_estilos_excel(ws)
            (h_font, h_fill, border, c_align, l_align, r_align, sec_font, g_fill, r_fill, g_font, r_font, i_font) = styles_excel

            row_idx = 1
            ws.cell(row_idx, 1, "Informações Gerais do Treinamento").font = sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_col_order))
            row_idx += 2
            
            info = {
                "Nome do Treinamento": self.entries["Nome do treinamento"].get(), "Turma": self.entries["Turma"].get(),
                "Modalidade": self.tipo_curso_var.get(), "Data(s) Realização": ", ".join([d.strftime("%d/%m/%Y") for d in self.dates]) or "N/A",
                "Horário Início": self.entries["Horário de Início"].get(), "Horário Término": self.entries["Horário de Término"].get(),
                "Público Alvo": self.entries["Público"].get(), "Número da Lista (Ref.)": self.entries["Número da Lista"].get(),
                "Link Gravação (se houver)": self.link_video_entry.get() or "N/A", "CFC (Opcional)": self.entries["CFC (Opcional)"].get() or "N/A"
            }
            info_keys = list(info.keys())
            for i in range(0, len(info_keys), 2):
                ws.cell(row_idx, 1, info_keys[i] + ":").font = Font(bold=True); ws.cell(row_idx, 1).alignment=Alignment(horizontal="right")
                ws.cell(row_idx, 2, info.get(info_keys[i], "-")).alignment = l_align
                if i + 1 < len(info_keys):
                    ws.cell(row_idx, 5, info_keys[i+1] + ":").font = Font(bold=True); ws.cell(row_idx, 5).alignment=Alignment(horizontal="right")
                    ws.cell(row_idx, 6, info.get(info_keys[i+1], "-")).alignment = l_align
                row_idx += 1
            
            row_idx += 1
            ws.cell(row_idx, 1, "Detalhamento da Carga Horária (Base de Cálculo)").font = Font(bold=True, underline="single"); row_idx += 1
            ws.cell(row_idx, 1, "C.H. Bruta por Dia:").font=Font(bold=True); ws.cell(row_idx, 2, f"{daily_gross_hours:.2f} h")
            ws.cell(row_idx, 3, "C.H. Bruta Total:").font=Font(bold=True); ws.cell(row_idx, 4, f"{total_gross_hours:.2f} h")
            row_idx += 1
            ws.cell(row_idx, 1, "Desconto Intervalo por Dia:").font=Font(bold=True); ws.cell(row_idx, 2, f"{daily_interval_hours:.2f} h")
            ws.cell(row_idx, 3, "Desconto Intervalo Total:").font=Font(bold=True); ws.cell(row_idx, 4, f"{total_interval_hours:.2f} h")
            row_idx += 1
            ws.cell(row_idx, 1, "C.H. LÍQUIDA (Base 100%):").font=Font(bold=True, color="00B050"); ws.cell(row_idx, 2, f"{net_training_hours:.2f} h").font=Font(bold=True, color="00B050")
            row_idx += 1

            if instrutores_data:
                row_idx +=1
                ws.cell(row_idx, 1, "Instrutores e Cargas Horárias").font = Font(bold=True, underline="single"); row_idx += 1
                for instrutor in instrutores_data:
                    ws.cell(row_idx, 1, "E-mail:").font = Font(bold=True); ws.cell(row_idx, 2, instrutor['email'])
                    ws.cell(row_idx, 3, "CH:").font = Font(bold=True); ws.cell(row_idx, 4, instrutor['ch'])
                    row_idx += 1
            
            row_idx += 1
            ws.cell(row_idx, 1, "Detalhes de Presença dos Participantes").font = sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_col_order))
            row_idx += 2
            
            h_r = row_idx
            ws.append(final_col_order)
            for c_idx, header_name in enumerate(final_col_order, 1):
                cell = ws.cell(h_r, c_idx); cell.font = h_font; cell.fill = h_fill; cell.border = border; cell.alignment = c_align
            
            for _, r_data in merged_df.iterrows():
                ws.append([r_data.get(c, "") for c in final_col_order])
                data_r = ws.max_row
                for c_idx, c_name in enumerate(final_col_order, 1):
                    cell = ws.cell(row=data_r, column=c_idx)
                    cell.border = border
                    if c_name in ['Nome', 'Email', 'CPF', 'Cargo', 'BU', 'Inscrição']: cell.alignment = l_align
                    elif c_name == 'Duracao_Total_Horas': cell.alignment, cell.number_format, cell.value = c_align, '0.00 "h"', float(cell.value or 0)
                    elif c_name == '% Presença': cell.alignment, cell.number_format, cell.value = c_align, '0.0"%"', float(cell.value or 0)
                    elif c_name == 'Status':
                        cell.alignment = c_align
                        if cell.value == "Presente": cell.fill, cell.font = g_fill, g_font
                        else: cell.fill, cell.font = r_fill, r_font
                    elif c_name == 'Observação':
                        cell.alignment = c_align
                        if cell.value == "Instrutor":
                            cell.font = Font(bold=True, color="0000FF"); cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            row_idx = ws.max_row + 2
            ws.cell(row_idx, 1, "Feedback NPS").font=sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_col_order))
            row_idx += 1
            ws.cell(row_idx, 1, "Score NPS:").font=Font(bold=True); ws.cell(row_idx, 2, nps_score_str)
            row_idx += 1
            ws.cell(row_idx, 1, "Detalhes Score:").font=Font(bold=True); ws.cell(row_idx, 2, nps_details_str)
            row_idx += 2
            ws.cell(row_idx, 1, "Feedback Consolidado (Texto Extraído):").font=Font(bold=True)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(final_col_order))
            row_idx += 1
            fb_cell=ws.cell(row_idx, 1, nps_texto_completo or "N/A")
            fb_cell.alignment=Alignment(wrap_text=True, vertical="top"); fb_cell.border=border
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 9, end_column=len(final_col_order))
            row_idx += 10
            
            ws.cell(row_idx, 1, f"Gerado por: {self.nome_usuario} em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font=i_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_col_order))
            ws.cell(row_idx, 1).alignment = Alignment(horizontal="left")

            # <<<<<<< NOVA FUNCIONALIDADE: Passa o flag para a função de divergências >>>>>>>
            self._criar_aba_divergencias(wb, nao_presentes, extras_na_presenca, is_master_list_mode=self.use_professionals_report_var.get())
            # --- FIM DA NOVA FUNCIONALIDADE ---

            self.progress.set(0.8)
            self.root.update_idletasks()

            print("\n7. Salvando Arquivo Excel (Digital)...")
            nome_safe = re.sub(r'[\\/*?:"<>|]', "", self.entries["Nome do treinamento"].get()).replace(' ', '_')[:50]
            data_safe = self.dates[0].strftime("%Y%m%d") if self.dates else "semdata"
            lp_safe = f"_LP{self.entries['Número da Lista'].get()}" if self.entries['Número da Lista'].get() else ""
            suggested_filename = f"Relatorio_Presenca_{nome_safe}{lp_safe}_{data_safe}_Digital.xlsx"
            
            f_path_save = filedialog.asksaveasfilename(initialfile=suggested_filename, title="Salvar Relatório Digital Como...", defaultextension=".xlsx", filetypes=[("Planilhas Excel", "*.xlsx")], parent=self.root)
            if not f_path_save:
                messagebox.showinfo("Cancelado", "Salvamento cancelado.", parent=self.root)
                self.progress.stop(); self.progress.set(0); self.gerar_button.configure(state=tk.NORMAL)
                return
            
            wb.save(f_path_save)
            messagebox.showinfo("Sucesso", f"Relatório Digital salvo em:\n{f_path_save}", parent=self.root)
            self.progress.set(0.9)
            self.root.update_idletasks()
            
            if self.emails_entry.get().strip():
                print("\n8. Enviando e-mail...")
                self.enviar_email(f_path_save)
            else:
                self.progress.set(1.0)
            
            print("\n--- Processo Digital Concluído ---")

        except Exception as e:
            error_details = traceback.format_exc()
            messagebox.showerror("Erro Fatal (Digital)", f"Erro:\n{type(e).__name__}: {e}\n\nDetalhes no console.", parent=self.root)
            print(f"ERRO INESPERADO (Digital):\n{error_details}")

    def gerar_relatorio_hibrido(self):
        try:
            print("\n--- Iniciando Geração Relatório Híbrido ---")
            self.progress.set(0.05)
            self.root.update_idletasks()

            # <<<<<<< NOVA FUNCIONALIDADE: Usa a nova função para obter a base de participantes >>>>>>>
            print("1. Obtendo base de participantes...")
            convidados_df = self._get_base_participants_df()
            # --- FIM DA NOVA FUNCIONALIDADE ---

            email_col = find_col_ignore_case(convidados_df.columns, ['email', 'e-mail'])
            if not email_col: raise ValueError("Modo Híbrido requer coluna 'Email' na lista de participantes (convidados ou profissionais).")
            convidados_df.rename(columns={email_col:'Email'}, inplace=True)
            convidados_df['Email'] = convidados_df['Email'].astype(str).str.lower().str.strip()
            
            cols_cv_para_merge = ['Email', 'Nome', 'Inscrição']
            nome_col_conv_orig = find_col_ignore_case(convidados_df.columns, ['nome'])
            if nome_col_conv_orig and 'Nome' not in convidados_df.columns:
                convidados_df.rename(columns={nome_col_conv_orig: 'Nome'}, inplace=True)
            if 'Nome' not in convidados_df.columns:
                convidados_df['Nome'] = ""

            map_conv = {'CPF':['cpf'],'Cargo':['cargo', 'função'],'BU':['bu', 'unidade'], 'Nome':['nome']}
            for std_name, names in map_conv.items():
                found=find_col_ignore_case(convidados_df.columns, names)
                if found and std_name not in convidados_df.columns: convidados_df.rename(columns={found:std_name}, inplace=True)
                if std_name in convidados_df.columns and std_name not in cols_cv_para_merge: cols_cv_para_merge.append(std_name)

            convidados_final_df = convidados_df.drop_duplicates(subset=['Email'])[cols_cv_para_merge].copy()

            print("   > Obtendo dados dos instrutores...")
            instrutores_data, instrutores_emails = self._get_instrutores_data()
            instrutores_emails_set = set(instrutores_emails)
            if instrutores_emails_set:
                print(f"     - Instrutores a serem sinalizados (Híbrido): {instrutores_emails_set}")
            
            self.progress.set(0.15)
            self.root.update_idletasks()

            print("\n2. Processando arquivos de presença (Modo Híbrido)...")
            participantes_sessoes = collections.Counter()
            participantes_duracao = collections.defaultdict(float)
            arquivos_falha = []
            listas_presenciais_count = 0

            mapa_nome_email = {}
            if 'Nome' in convidados_df.columns and 'Email' in convidados_df.columns:
                convidados_com_nome_email = convidados_df[convidados_df['Nome'].notna() & (convidados_df['Nome'] != '') & convidados_df['Email'].notna()].copy()
                convidados_com_nome_email['Nome_lower_strip'] = convidados_com_nome_email['Nome'].astype(str).str.lower().str.strip()
                mapa_nome_email = pd.Series(convidados_com_nome_email.Email.values, index=convidados_com_nome_email.Nome_lower_strip).to_dict()
                if not mapa_nome_email:
                    print("   AVISO (Híbrido): Mapa Nome->Email está vazio. Listas presenciais baseadas em nome podem não ser processadas.")
            else:
                print("   AVISO (Híbrido): Coluna 'Nome' ou 'Email' ausente nos participantes. Mapeamento Nome->Email desabilitado.")

            
            for file_path in self.presenca_paths:
                tipo, dados_retornados = self._processar_arquivo_presenca(file_path, id_col_primaria='Email')
                                
                if tipo == 'presencial':
                    emails_para_contagem_sessao = set()
                    for ident_bruto in dados_retornados:
                        ident_str = str(ident_bruto).lower().strip()
                        if '@' in ident_str:
                            emails_para_contagem_sessao.add(ident_str)
                        elif ident_str in mapa_nome_email:
                            emails_para_contagem_sessao.add(mapa_nome_email[ident_str])
                    
                    if emails_para_contagem_sessao:
                        participantes_sessoes.update(emails_para_contagem_sessao)
                        listas_presenciais_count += 1
                    elif dados_retornados:
                        arquivos_falha.append(f"{os.path.basename(file_path)} (Presencial, mas sem e-mails ou nomes mapeáveis)")
                        print(f"      AVISO (Híbrido): Arquivo presencial {os.path.basename(file_path)} não continha e-mails ou nomes mapeáveis.")
                elif tipo == 'digital':
                    for _, row in dados_retornados.iterrows():
                        participantes_duracao[row['Email'].lower()] += row['Duracao_horas']
                else:
                    arquivos_falha.append(dados_retornados)
            
            if not participantes_sessoes and not participantes_duracao:
                raise ValueError(f"Nenhum dado de presença válido extraído.\nErros: {arquivos_falha}")
            
            self.progress.set(0.4)
            self.root.update_idletasks()
            
            print("\n3. Consolidando dados de presença...")
            df_sessoes = pd.DataFrame(participantes_sessoes.items(), columns=['Email', 'Sessoes_Presente'])
            df_duracao = pd.DataFrame(participantes_duracao.items(), columns=['Email', 'Duracao_Total_Horas'])
            
            if df_sessoes.empty and df_duracao.empty:
                presenca_agregada = pd.DataFrame(columns=['Email', 'Sessoes_Presente', 'Duracao_Total_Horas'])
            elif df_sessoes.empty:
                presenca_agregada = df_duracao.copy()
            elif df_duracao.empty:
                presenca_agregada = df_sessoes.copy()
            else:
                presenca_agregada = pd.merge(df_sessoes, df_duracao, on='Email', how='outer')

            if 'Sessoes_Presente' not in presenca_agregada.columns: presenca_agregada['Sessoes_Presente'] = 0
            else: presenca_agregada['Sessoes_Presente'].fillna(0, inplace=True)
            
            if 'Duracao_Total_Horas' not in presenca_agregada.columns: presenca_agregada['Duracao_Total_Horas'] = 0.0
            else: presenca_agregada['Duracao_Total_Horas'].fillna(0.0, inplace=True)
            
            presenca_agregada['Sessoes_Presente'] = presenca_agregada['Sessoes_Presente'].astype(int)
            presenca_agregada['Duracao_Total_Horas'] = presenca_agregada['Duracao_Total_Horas'].astype(float)

            net_hours, total_gross_hours, total_interval_hours, daily_gross_hours, daily_interval_hours = self._get_effective_training_hours()
            
            presenca_agregada['% Presença Presencial'] = 0.0
            if listas_presenciais_count > 0:
                presenca_agregada['% Presença Presencial'] = (presenca_agregada['Sessoes_Presente'] / listas_presenciais_count * 100).clip(upper=100)
            
            print(f"   > Deduzindo o tempo total de intervalo ({total_interval_hours:.2f}h) da duração digital dos participantes.")
            if 'Duracao_Total_Horas' in presenca_agregada.columns:
                presenca_agregada['Duracao_Total_Horas'] = (presenca_agregada['Duracao_Total_Horas'] - total_interval_hours).clip(lower=0)

            presenca_agregada['% Presença Digital'] = 0.0
            if net_hours > 0:
                presenca_agregada['% Presença Digital'] = (presenca_agregada['Duracao_Total_Horas'] / net_hours * 100).clip(upper=100)

            print("\n4. Combinando dados e preparando o relatório final...")
            # <<<<<<< NOVA FUNCIONALIDADE: Lógica de merge condicional >>>>>>>
            merge_how = "inner" if self.use_professionals_report_var.get() else "left"
            print(f"   > Usando modo de merge: '{merge_how}'.")
            merged = pd.merge(convidados_final_df, presenca_agregada, on='Email', how=merge_how)
            # --- FIM DA NOVA FUNCIONALIDADE ---
            
            cols_from_presenca = ['Sessoes_Presente', 'Duracao_Total_Horas', '% Presença Presencial', '% Presença Digital']
            for col_p in cols_from_presenca:
                if col_p in merged.columns: merged[col_p].fillna(0, inplace=True)
                else: merged[col_p] = 0

            presente_presencial = merged['% Presença Presencial'] >= 75
            presente_digital = merged['% Presença Digital'] >= 75
            merged['Status'] = "Falta"
            merged.loc[presente_presencial | presente_digital, 'Status'] = "Presente"

            merged['Observação'] = ''
            if instrutores_emails_set:
                merged.loc[merged['Email'].isin(instrutores_emails_set), 'Observação'] = 'Instrutor'

            final_cols_order = ['Nome', 'Email', 'CPF', 'Status', 'Sessoes_Presente', 'Duracao_Total_Horas', '% Presença Presencial', '% Presença Digital', 'Inscrição', 'Cargo', 'BU', 'Observação']
            for col in final_cols_order:
                if col not in merged.columns: merged[col] = ""

            if 'Nome' not in merged.columns: merged['Nome'] = ""
            merged['Nome'].fillna("", inplace=True)
            merged = merged[final_cols_order].sort_values(by="Nome", key=lambda c: c.astype(str).str.lower())
            
            self.progress.set(0.6)
            self.root.update_idletasks()
            
            print("\n5. Verificando Divergências...")
            temp_presenca_agg = pd.DataFrame({'Email': list(set(participantes_sessoes.keys()).union(set(participantes_duracao.keys())))})
            nao_presentes, extras_na_presenca = self.verificar_divergencias(convidados_df, temp_presenca_agg)

            print("\n6. Processando NPS...")
            nps_path = self.nps_path.get()
            nps_score_str, nps_details_str, nps_texto_completo = self.processar_nps(nps_path) if nps_path and os.path.exists(nps_path) else ("N/A", "Arquivo não fornecido.", "")
            
            print("\n7. Gerando Arquivo Excel (Híbrido)...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório Presença Híbrido"
            styles_excel = self._aplicar_estilos_excel(ws)
            (h_font, h_fill, border, c_align, l_align, r_align, sec_font, g_fill, r_fill, g_font, r_font, i_font) = styles_excel
            
            for col_letter, width in zip("ABCDEFGHIJKL", [35, 35, 18, 12, 18, 18, 22, 22, 15, 25, 20, 15]):
                ws.column_dimensions[col_letter].width = width

            row_idx = 1
            ws.cell(row_idx, 1, "Informações Gerais do Treinamento").font = sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order))
            row_idx += 2
            
            info = {
                "Nome do Treinamento": self.entries["Nome do treinamento"].get(), "Turma": self.entries["Turma"].get(),
                "Modalidade": self.tipo_curso_var.get(), "Data(s) Realização": ", ".join([d.strftime("%d/%m/%Y") for d in self.dates]) or "N/A",
                "Horário Início": self.entries["Horário de Início"].get() or "N/A", "Horário Término": self.entries["Horário de Término"].get() or "N/A",
                "Público Alvo": self.entries["Público"].get(), "Número da Lista (Ref.)": self.entries["Número da Lista"].get(),
                "Link Gravação (se houver)": self.link_video_entry.get() or "N/A", "CFC (Opcional)": self.entries["CFC (Opcional)"].get() or "N/A"
            }

            info_keys = list(info.keys())
            for i in range(0, len(info_keys), 2):
                ws.cell(row_idx, 1, info_keys[i] + ":").font = Font(bold=True); ws.cell(row_idx, 1).alignment=Alignment(horizontal="right")
                ws.cell(row_idx, 2, info.get(info_keys[i], "-")).alignment = l_align
                if i + 1 < len(info_keys):
                    ws.cell(row_idx, 5, info_keys[i+1] + ":").font = Font(bold=True); ws.cell(row_idx, 5).alignment=Alignment(horizontal="right")
                    ws.cell(row_idx, 6, info.get(info_keys[i+1], "-")).alignment = l_align
                row_idx += 1
            
            row_idx += 1
            ws.cell(row_idx, 1, "Detalhamento da Carga Horária (Base de Cálculo)").font = Font(bold=True, underline="single"); row_idx += 1
            ws.cell(row_idx, 1, "C.H. Bruta por Dia:").font=Font(bold=True); ws.cell(row_idx, 2, f"{daily_gross_hours:.2f} h")
            ws.cell(row_idx, 3, "C.H. Bruta Total:").font=Font(bold=True); ws.cell(row_idx, 4, f"{total_gross_hours:.2f} h")
            row_idx += 1
            ws.cell(row_idx, 1, "Desconto Intervalo por Dia:").font=Font(bold=True); ws.cell(row_idx, 2, f"{daily_interval_hours:.2f} h")
            ws.cell(row_idx, 3, "Desconto Intervalo Total:").font=Font(bold=True); ws.cell(row_idx, 4, f"{total_interval_hours:.2f} h")
            row_idx += 1
            ws.cell(row_idx, 1, "C.H. LÍQUIDA (Base 100%):").font=Font(bold=True, color="00B050"); ws.cell(row_idx, 2, f"{net_hours:.2f} h").font=Font(bold=True, color="00B050")
            ws.cell(row_idx, 5, "Total Listas Presenciais:").font=Font(bold=True); ws.cell(row_idx, 6, listas_presenciais_count)
            row_idx += 1

            if instrutores_data:
                row_idx +=1
                ws.cell(row_idx, 1, "Instrutores e Cargas Horárias").font = Font(bold=True, underline="single"); row_idx += 1
                for instrutor in instrutores_data:
                    ws.cell(row_idx, 1, "E-mail:").font = Font(bold=True); ws.cell(row_idx, 2, instrutor['email'])
                    ws.cell(row_idx, 3, "CH:").font = Font(bold=True); ws.cell(row_idx, 4, instrutor['ch'])
                    row_idx += 1

            row_idx += 1
            ws.cell(row_idx, 1, "Participantes e Status (Híbrido)").font = sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order))
            row_idx += 2
            
            h_r = row_idx
            ws.append(final_cols_order)
            for c_idx, header_name in enumerate(final_cols_order, 1):
                cell = ws.cell(h_r, c_idx); cell.font = h_font; cell.fill = h_fill; cell.border = border; cell.alignment = c_align
            
            for _, r_data in merged.iterrows():
                ws.append([r_data.get(c, "") for c in final_cols_order])
                data_r = ws.max_row
                for c_idx, c_name in enumerate(final_cols_order, 1):
                    cell = ws.cell(row=data_r, column=c_idx)
                    cell.border = border
                    
                    if c_name in ['Nome', 'Email', 'CPF', 'Cargo', 'BU', 'Inscrição']: cell.alignment = l_align
                    elif c_name == 'Status':
                        cell.alignment = c_align
                        if cell.value == "Presente": cell.fill, cell.font = g_fill, g_font
                        else: cell.fill, cell.font = r_fill, r_font
                    elif c_name == 'Sessoes_Presente':
                        cell.alignment = c_align
                        cell.number_format = f'0" de {listas_presenciais_count}"'
                    elif c_name == 'Duracao_Total_Horas':
                        cell.alignment = c_align
                        cell.number_format = '0.00 "h"'
                    elif c_name in ['% Presença Presencial', '% Presença Digital']:
                        cell.alignment = c_align
                        cell.number_format = '0.0"%"'
                    elif c_name == 'Observação':
                        cell.alignment = c_align
                        if cell.value == "Instrutor":
                            cell.font = Font(bold=True, color="0000FF"); cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    else: cell.alignment = c_align

            row_idx = ws.max_row + 2
            ws.cell(row_idx, 1, "Feedback NPS").font=sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order))
            row_idx += 1
            ws.cell(row_idx, 1, "Score NPS:").font=Font(bold=True); ws.cell(row_idx, 2, nps_score_str)
            row_idx += 1
            ws.cell(row_idx, 1, "Detalhes Score:").font=Font(bold=True); ws.cell(row_idx, 2, nps_details_str)
            row_idx += 2
            ws.cell(row_idx, 1, "Feedback Consolidado (Texto Extraído):").font=Font(bold=True)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(final_cols_order))
            row_idx += 1
            fb_cell=ws.cell(row_idx, 1, nps_texto_completo or "N/A")
            fb_cell.alignment=Alignment(wrap_text=True, vertical="top"); fb_cell.border=border
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 9, end_column=len(final_cols_order))
            row_idx += 10
            
            ws.cell(row_idx, 1, f"Gerado por: {self.nome_usuario} em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font=i_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order))
            ws.cell(row_idx, 1).alignment = Alignment(horizontal="left")

            # <<<<<<< NOVA FUNCIONALIDADE: Passa o flag para a função de divergências >>>>>>>
            self._criar_aba_divergencias(wb, nao_presentes, extras_na_presenca, is_master_list_mode=self.use_professionals_report_var.get())
            # --- FIM DA NOVA FUNCIONALIDADE ---

            self.progress.set(0.8)
            self.root.update_idletasks()
            
            print("\n8. Salvando Arquivo Excel (Híbrido)...")
            nome_safe = re.sub(r'[\\/*?:"<>|]', "", self.entries["Nome do treinamento"].get()).replace(' ', '_')[:50]
            data_safe = self.dates[0].strftime("%Y%m%d") if self.dates else "semdata"
            lp_safe = f"_LP{self.entries['Número da Lista'].get()}" if self.entries['Número da Lista'].get() else ""
            suggested_filename = f"Relatorio_Presenca_{nome_safe}{lp_safe}_{data_safe}_Hibrido.xlsx"
            f_path_save = filedialog.asksaveasfilename(initialfile=suggested_filename, title="Salvar Relatório Híbrido Como...", defaultextension=".xlsx", filetypes=[("Planilhas Excel", "*.xlsx")], parent=self.root)

            if not f_path_save:
                messagebox.showinfo("Cancelado", "Salvamento cancelado.", parent=self.root)
                self.progress.stop(); self.progress.set(0); self.gerar_button.configure(state=tk.NORMAL)
                return
            
            wb.save(f_path_save)
            messagebox.showinfo("Sucesso", f"Relatório Híbrido salvo em:\n{f_path_save}", parent=self.root)
            self.progress.set(0.9)
            self.root.update_idletasks()
            
            if self.emails_entry.get().strip():
                print("\n9. Enviando e-mail...")
                self.enviar_email(f_path_save)
            else:
                self.progress.set(1.0)
            
            print("\n--- Processo Híbrido Concluído ---")

        except Exception as e:
            error_details = traceback.format_exc()
            messagebox.showerror("Erro Fatal (Híbrido)", f"Erro:\n{type(e).__name__}: {e}\n\nDetalhes no console.", parent=self.root)
            print(f"ERRO INESPERADO (Híbrido):\n{error_details}")

    def gerar_relatorio_presencial(self):
        print("\n--- Iniciando Geração Relatório Presencial (Lógica de Contagem e Priorização) ---")
        try:
            self.progress.set(0.05)
            self.root.update_idletasks()
            
            # <<<<<<< NOVA FUNCIONALIDADE: Usa a nova função para obter a base de participantes >>>>>>>
            print("1. Obtendo base de participantes...")
            convidados_df = self._get_base_participants_df()
            # --- FIM DA NOVA FUNCIONALIDADE ---
            
            email_col=find_col_ignore_case(convidados_df.columns, ['email','e-mail'])
            nome_col=find_col_ignore_case(convidados_df.columns, ['nome','name', 'nome completo'])
            if not email_col and not nome_col:
                raise ValueError("A base de participantes deve conter coluna 'Email' ou 'Nome'.")
            
            id_col_primario = 'Email' if email_col else 'Nome'
            
            if email_col and email_col != 'Email': convidados_df.rename(columns={email_col: 'Email'}, inplace=True)
            if nome_col and nome_col != 'Nome': convidados_df.rename(columns={nome_col: 'Nome'}, inplace=True)
            
            if 'Email' in convidados_df.columns: convidados_df['Email'] = convidados_df['Email'].astype(str).str.lower().str.strip()
            if 'Nome' in convidados_df.columns: convidados_df['Nome'] = convidados_df['Nome'].astype(str).str.strip()
            
            if id_col_primario == 'Email':
                convidados_df.dropna(subset=['Email'], inplace=True)
                convidados_df = convidados_df[convidados_df['Email'].str.contains('@', na=False)]
            else:
                convidados_df.dropna(subset=['Nome'], inplace=True)
                convidados_df = convidados_df[convidados_df['Nome'] != '']
            
            if convidados_df.empty:
                raise ValueError(f"Nenhum participante válido encontrado com ID '{id_col_primario}'.")
            
            print(f"   > Base de participantes lida ({len(convidados_df)}). ID primário para busca: '{id_col_primario}'.")
            
            print("   > Obtendo dados dos instrutores...")
            instrutores_data, instrutores_emails = self._get_instrutores_data()
            instrutores_emails_set = set(instrutores_emails)
            if instrutores_emails_set:
                print(f"     - Instrutores a serem sinalizados (Presencial): {instrutores_emails_set}")

            self.progress.set(0.15)
            self.root.update_idletasks()

            print(f"\n2. Processando Arquivos de Presença e contando ocorrências...")
            total_listas_presenca = len(self.presenca_paths)
            if total_listas_presenca == 0:
                raise ValueError("Nenhuma lista de presença foi anexada.")

            presentes_contador = collections.Counter()
            arquivos_falha = []

            for i, f_path in enumerate(self.presenca_paths):
                fname = os.path.basename(f_path)
                print(f"   > Processando lista {i+1}/{total_listas_presenca}: {fname}")
                try:
                    p_df = self._read_input_file(f_path)
                    
                    id_pres = None
                    colunas_prioritarias_email = ["E-mail KPMG", "E-mail Corporativo", "Email", "e-mail"]
                    colunas_prioritarias_nome = ["Nome Completo", "Name", "Nome"]
                    
                    opcoes_busca = colunas_prioritarias_email + colunas_prioritarias_nome
                    if id_col_primario == 'Nome':
                        opcoes_busca = colunas_prioritarias_nome + colunas_prioritarias_email
                    
                    for col_opt in opcoes_busca:
                        id_pres = find_col_ignore_case(p_df.columns, [col_opt])
                        if id_pres:
                            print(f"      > Usando coluna prioritária '{id_pres}' como ID neste arquivo.")
                            break

                    if not id_pres:
                        raise ValueError(f"Nenhuma coluna de ID utilizável encontrada em '{fname}'.")

                    ids_in_file = set(p_df[id_pres].dropna().astype(str).str.strip())
                    
                    is_email_col_pres = any(keyword in id_pres.lower() for keyword in colunas_prioritarias_email)
                    if is_email_col_pres:
                        ids_in_file = {s.lower() for s in ids_in_file if '@' in s}
                    else:
                        ids_in_file.discard('')
                    
                    presentes_contador.update(ids_in_file)
                    print(f"      > {len(ids_in_file)} IDs únicos encontrados. Contagens atualizadas.")

                except Exception as e:
                    arquivos_falha.append(f"{fname} (Erro: {e})")

            if not presentes_contador:
                raise ValueError("Nenhum participante presente foi encontrado nos arquivos processados.")
            
            self.progress.set(0.40)
            self.root.update_idletasks()
            
            print("\n3. Combinando Dados e Calculando Percentual de Presença...")
            presenca_df = pd.DataFrame(presentes_contador.items(), columns=[id_col_primario, 'Contagem_Presenca'])
            presenca_df['% Presença'] = (presenca_df['Contagem_Presenca'] / total_listas_presenca * 100).clip(upper=100)
            
            cols_cv_para_merge = [id_col_primario, 'Nome', 'Inscrição']
            if 'Email' in convidados_df.columns: cols_cv_para_merge.append('Email')
            for col_padrao in ['Cargo','BU','CPF']:
                col_encontrada = find_col_ignore_case(convidados_df.columns,[col_padrao.lower(), col_padrao])
                if col_encontrada and col_encontrada != col_padrao and col_padrao not in convidados_df.columns:
                    convidados_df.rename(columns={col_encontrada:col_padrao}, inplace=True)
                if col_padrao in convidados_df.columns and col_padrao not in cols_cv_para_merge:
                    cols_cv_para_merge.append(col_padrao)
            
            # <<<<<<< NOVA FUNCIONALIDADE: Lógica de merge condicional >>>>>>>
            merge_how = "inner" if self.use_professionals_report_var.get() else "left"
            print(f"   > Usando modo de merge: '{merge_how}'.")
            merged = pd.merge(convidados_df.drop_duplicates(subset=[id_col_primario])[list(set(cols_cv_para_merge))], presenca_df, on=id_col_primario, how=merge_how)
            # --- FIM DA NOVA FUNCIONALIDADE ---
            
            merged['Contagem_Presenca'].fillna(0, inplace=True); merged['% Presença'].fillna(0, inplace=True)
            merged['Contagem_Presenca'] = merged['Contagem_Presenca'].astype(int)
            merged['Status'] = merged['% Presença'].apply(lambda x: "Presente" if x >= 75 else "Falta")
            
            merged['Observação'] = ''
            if instrutores_emails_set and 'Email' in merged.columns:
                merged.loc[merged['Email'].isin(instrutores_emails_set), 'Observação'] = 'Instrutor'
            elif instrutores_emails_set:
                print("   <!> AVISO (Presencial): Impossível marcar instrutores pois coluna 'Email' não está no resultado final.")

            final_cols_order_pres = ['Nome','Email','CPF', 'Contagem_Presenca', '% Presença', 'Status', 'Inscrição', 'Cargo','BU', 'Observação']
            for col in final_cols_order_pres:
                if col not in merged.columns:
                    merged[col] = ""
            
            merged = merged[final_cols_order_pres].sort_values(by='Nome', key=lambda c: c.astype(str).str.lower())
            
            self.progress.set(0.60)
            self.root.update_idletasks()
            
            print("\n4. Verificando Divergências...")
            emails_com_alguma_presenca = merged[merged['Contagem_Presenca'] > 0]['Email'].dropna().unique()
            presenca_para_verificacao = pd.DataFrame({'Email': emails_com_alguma_presenca})

            nao_presentes, extras_na_presenca = self.verificar_divergencias(convidados_df, presenca_para_verificacao)

            print("\n5. Processando NPS...")
            nps_path = self.nps_path.get()
            nps_score_str, nps_details_str, nps_texto_completo = self.processar_nps(nps_path) if nps_path and os.path.exists(nps_path) else ("N/A", "Arquivo não fornecido.", "")
            
            print("\n6. Gerando Arquivo Excel (Presencial)...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório Presencial"
            styles_excel = self._aplicar_estilos_excel(ws)
            (h_font, h_fill, border, c_align, l_align, r_align, sec_font, g_fill, r_fill, g_font, r_font, i_font) = styles_excel
            
            row_idx = 1
            ws.cell(row_idx, 1, "Informações Gerais do Treinamento").font=sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order_pres))
            row_idx += 2
            
            info = {
                "Nome do Treinamento": self.entries["Nome do treinamento"].get(), "Turma": self.entries["Turma"].get(),
                "Modalidade": self.tipo_curso_var.get(), "Data(s) Realização": ", ".join([d.strftime("%d/%m/%Y") for d in self.dates]) or "N/A",
                "Horário Início": self.entries["Horário de Início"].get() or "N/A", "Horário Término": self.entries["Horário de Término"].get() or "N/A",
                "Carga Horária Bruta Total (h)": self.entries["Carga Horária Bruta Total (h)"].get() or "N/A", "Total de Sessões": total_listas_presenca,
                "Público Alvo": self.entries["Público"].get(), "Número da Lista (Ref.)": self.entries["Número da Lista"].get(),
                "Link Gravação (se houver)": self.link_video_entry.get() or "N/A",
                "CFC (Opcional)": self.entries["CFC (Opcional)"].get() or "N/A"
            }
            info_keys = list(info.keys())
            for i in range(0, len(info_keys), 2):
                ws.cell(row_idx, 1, info_keys[i] + ":").font = Font(bold=True); ws.cell(row_idx, 1).alignment=Alignment(horizontal="right")
                ws.cell(row_idx, 2, info.get(info_keys[i], "-")).alignment = l_align
                if i + 1 < len(info_keys):
                    ws.cell(row_idx, 5, info_keys[i+1] + ":").font = Font(bold=True); ws.cell(row_idx, 5).alignment=Alignment(horizontal="right")
                    ws.cell(row_idx, 6, info.get(info_keys[i+1], "-")).alignment = l_align
                row_idx += 1
                
            if instrutores_data:
                row_idx +=1
                ws.cell(row_idx, 1, "Instrutores e Cargas Horárias").font = Font(bold=True, underline="single"); row_idx += 1
                for instrutor in instrutores_data:
                    ws.cell(row_idx, 1, "E-mail:").font = Font(bold=True); ws.cell(row_idx, 2, instrutor['email'])
                    ws.cell(row_idx, 3, "CH:").font = Font(bold=True); ws.cell(row_idx, 4, instrutor['ch'])
                    row_idx += 1

            row_idx +=1
            ws.cell(row_idx, 1, "Lista de Participantes (Presencial)").font=sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order_pres))
            row_idx += 1
            ws.cell(row_idx, 1, f"Status baseado na presença em {total_listas_presenca} sessões/listas.").font=Font(italic=True, size=9)
            row_idx += 1
            
            h_r = row_idx
            ws.append(final_cols_order_pres)
            for c_idx, header_name in enumerate(final_cols_order_pres, 1):
                cell=ws.cell(h_r, c_idx); cell.font=h_font; cell.fill=h_fill; cell.border=border; cell.alignment=c_align

            for _, r_data in merged.iterrows():
                ws.append([r_data.get(c, "") for c in final_cols_order_pres])
                data_r = ws.max_row
                for c_idx, c_name in enumerate(final_cols_order_pres, 1):
                    cell = ws.cell(row=data_r, column=c_idx)
                    cell.border=border
                    if c_name in ['Nome', 'Email', 'CPF', 'Cargo', 'BU', 'Inscrição']: cell.alignment = l_align
                    elif c_name == 'Contagem_Presenca': cell.alignment, cell.number_format = c_align, f'0" de {total_listas_presenca}"'
                    elif c_name == '% Presença': cell.alignment, cell.number_format, cell.value = c_align, '0.0"%"', float(cell.value or 0)
                    elif c_name == 'Status':
                        cell.alignment = c_align
                        if cell.value == "Presente": cell.fill, cell.font = g_fill, g_font
                        else: cell.fill, cell.font = r_fill, r_font
                    elif c_name == 'Observação':
                        cell.alignment = c_align
                        if cell.value == "Instrutor":
                            cell.font = Font(bold=True, color="0000FF"); cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            row_idx = ws.max_row + 2
            ws.cell(row_idx, 1, "Feedback NPS").font=sec_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order_pres))
            row_idx += 1
            ws.cell(row_idx, 1, "Score NPS:").font=Font(bold=True); ws.cell(row_idx, 2, nps_score_str)
            row_idx += 1
            ws.cell(row_idx, 1, "Detalhes Score:").font=Font(bold=True); ws.cell(row_idx, 2, nps_details_str)
            row_idx += 2
            ws.cell(row_idx, 1, "Feedback Consolidado (Texto Extraído):").font=Font(bold=True)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(final_cols_order_pres))
            row_idx += 1
            fb_cell=ws.cell(row_idx, 1, nps_texto_completo or "N/A")
            fb_cell.alignment=Alignment(wrap_text=True, vertical="top"); fb_cell.border=border
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 9, end_column=len(final_cols_order_pres))
            row_idx += 10
            
            ws.cell(row_idx, 1, f"Gerado por: {self.nome_usuario} em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font=i_font
            ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(final_cols_order_pres))
            ws.cell(row_idx, 1).alignment = Alignment(horizontal="left")

            # <<<<<<< NOVA FUNCIONALIDADE: Passa o flag para a função de divergências >>>>>>>
            self._criar_aba_divergencias(wb, nao_presentes, extras_na_presenca, is_master_list_mode=self.use_professionals_report_var.get())
            # --- FIM DA NOVA FUNCIONALIDADE ---

            self.progress.set(0.8)
            self.root.update_idletasks()
            
            print("\n7. Salvando Arquivo Excel (Presencial)...")
            nome_safe = re.sub(r'[\\/*?:"<>|]', "", self.entries["Nome do treinamento"].get()).replace(' ', '_')[:50]
            data_safe = self.dates[0].strftime("%Y%m%d") if self.dates else "semdata"
            lp_safe = f"_LP{self.entries['Número da Lista'].get()}" if self.entries['Número da Lista'].get() else ""
            suggested_filename = f"Relatorio_Presenca_{nome_safe}{lp_safe}_{data_safe}_Presencial.xlsx"
            
            f_path_save = filedialog.asksaveasfilename(initialfile=suggested_filename, title="Salvar Relatório Presencial Como...", defaultextension=".xlsx", filetypes=[("Planilhas Excel", "*.xlsx")], parent=self.root)

            if not f_path_save:
                messagebox.showinfo("Cancelado", "Salvamento cancelado.", parent=self.root)
                self.progress.stop(); self.progress.set(0); self.gerar_button.configure(state=tk.NORMAL)
                return
            
            wb.save(f_path_save)
            messagebox.showinfo("Sucesso", f"Relatório Presencial salvo em:\n{f_path_save}", parent=self.root)
            self.progress.set(0.9)
            self.root.update_idletasks()
            
            if self.emails_entry.get().strip():
                print("\n8. Enviando e-mail...")
                self.enviar_email(f_path_save)
            else:
                self.progress.set(1.0)
                
            print("\n--- Processo Presencial Concluído ---")

        except Exception as e:
            error_details = traceback.format_exc()
            messagebox.showerror("Erro Inesperado (Presencial)", f"Erro:\n{type(e).__name__}: {e}\n\nDetalhes no console.", parent=self.root)
            print(f"ERRO INESPERADO (Presencial):\n{error_details}")

    def enviar_email(self, file_path):
        emails_str = self.emails_entry.get().strip()
        if not emails_str:
            print("   > Nenhum e-mail fornecido para envio.")
            self.progress.set(1.0)
            return
        emails_list = [e.strip() for e in emails_str.replace(';', ',').split(',') if e.strip()]
        valid_emails = [e for e in emails_list if '@' in e and '.' in e.split('@')[-1]]
        if not valid_emails:
            messagebox.showwarning("E-mails Inválidos", "Nenhum endereço de e-mail válido encontrado para envio.", parent=self.root)
            print(f"   <!> Nenhum e-mail válido detectado em: '{emails_str}'")
            self.progress.set(1.0)
            return
        try:
            print(f"   > Preparando e-mail para: {'; '.join(valid_emails)}")
            try:
                outlook = win32.GetActiveObject('Outlook.Application')
                print("   > Instância ativa do Outlook encontrada.")
            except:
                print("   > Nenhuma instância ativa do Outlook. Iniciando uma nova...")
                outlook = win32.Dispatch('outlook.application')
                print("   > Nova instância do Outlook iniciada.")
            mail = outlook.CreateItem(0)
            mail.To = ";".join(valid_emails)
            t_nome = self.entries['Nome do treinamento'].get() or "Nome Não Informado"
            turma = self.entries['Turma'].get() or "Não Informada"
            numero_lista = self.entries['Número da Lista'].get() or "N/A"
            mod = self.tipo_curso_var.get()
            d_str = ", ".join([d.strftime('%d/%m/%Y') for d in self.dates]) if self.dates else 'Data Não Informada'
            
            mail.Subject = f"LP {numero_lista} - Relatório de Presença: {t_nome} ({mod})"

            mail.Body = (
                f"Prezados,\n\n"
                f"Segue em anexo o relatório de presença referente ao treinamento:\n\n"
                f"   - Nome: \"{t_nome}\"\n"
                f"   - Turma: \"{turma}\"\n"
                f"   - Modalidade: {mod}\n"
                f"   - Período(s): {d_str}\n"
                f"   - Número da Lista: {numero_lista}\n\n"
                f"Este relatório foi gerado automaticamente através da ferramenta interna.\n\n"
                f"Gerado por: {self.nome_usuario}\n\n"
                f"Atenciosamente,\n"
                f"Sistema de Geração de Relatórios"
            )
            abs_path = os.path.abspath(file_path)
            if not os.path.exists(abs_path):
                messagebox.showerror("Erro Anexo", f"Arquivo de relatório não encontrado para anexar:\n{abs_path}", parent=self.root)
                print(f"   <!> ERRO CRÍTICO: Anexo não encontrado no caminho absoluto: {abs_path}")
                self.progress.set(1.0)
                return
            mail.Attachments.Add(abs_path)
            print(f"   > Anexando arquivo: {abs_path}")
            mail.Send()
            print("   > Comando mail.Send() executado.")
            messagebox.showinfo("E-mail Enviado", f"E-mail com o relatório foi enviado com sucesso para:\n{', '.join(valid_emails)}", parent=self.root)
            self.progress.set(1.0)
        except AttributeError as e_attr:
            if "'NoneType' object has no attribute" in str(e_attr) or "invalid class string" in str(e_attr).lower():
                messagebox.showerror("Erro Outlook", "Não foi possível conectar ao Microsoft Outlook.\nVerifique se ele está instalado e configurado corretamente.", parent=self.root)
                print(f"ERRO DE COMUNICAÇÃO COM OUTLOOK: {e_attr}")
            else:
                error_details = traceback.format_exc()
                messagebox.showerror("Erro Atributo E-mail", f"Ocorreu um erro inesperado ao preparar o e-mail:\n{e_attr}", parent=self.root)
                print(f"ERRO EMAIL (AttributeError):\n{error_details}")
            self.progress.set(1.0)
        except Exception as e:
            error_details = traceback.format_exc()
            messagebox.showerror("Erro ao Enviar E-mail", f"Falha ao enviar o e-mail via Outlook:\n{type(e).__name__}: {e}", parent=self.root)
            print(f"ERRO ENVIO EMAIL:\n{error_details}")
            self.progress.set(1.0)

    def _log_to_email_tab(self, message):
        def _update_log_widget():
            if hasattr(self, 'log_envio_emails_text') and self.log_envio_emails_text.winfo_exists():
                self.log_envio_emails_text.configure(state=tk.NORMAL)
                self.log_envio_emails_text.insert(tk.END, f"{datetime.datetime.now().strftime('%H:%M:%S')} - {message}\n")
                self.log_envio_emails_text.see(tk.END)
                self.log_envio_emails_text.configure(state=tk.DISABLED)
            else:
                print(f"(LOG Console - Email Tab): {message}")

        if self.root.winfo_exists():
            self.root.after(0, _update_log_widget)

    def _enviar_emails_pos_treinamento_action(self):
        report_path = filedialog.askopenfilename(
            title="Selecione o Relatório de Presença Gerado (Excel)",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            parent=self.root
        )
        if not report_path:
            messagebox.showinfo("Cancelado", "Nenhum relatório selecionado.", parent=self.root)
            return

        self.enviar_pos_treinamento_button.configure(state=tk.DISABLED)
        self.progress_emails.set(0)
        self.root.update_idletasks()
        self.progress_emails.start()
        self.root.after(150, lambda: self._enviar_emails_pos_treinamento_async(report_path))

    def _enviar_emails_pos_treinamento_async(self, report_path):
        try:
            self._log_to_email_tab(f"Lendo relatório: {os.path.basename(report_path)}")
            try:
                xls = pd.ExcelFile(report_path)
                sheet_name_to_read = None
                common_sheet_names = ["Relatório Presença", "Relatório Presença Híbrido", "Relatório Presencial"]
                for name in common_sheet_names:
                    if name in xls.sheet_names:
                        sheet_name_to_read = name
                        break
                if not sheet_name_to_read and xls.sheet_names: sheet_name_to_read = xls.sheet_names[0]
                if not sheet_name_to_read: raise ValueError("Nenhuma aba encontrada no Excel.")
                df_temp = pd.read_excel(xls, sheet_name=sheet_name_to_read, header=None)
                
                header_actual_index = -1
                header_keywords = [
                    "Detalhes de Presença dos Participantes",
                    "Lista de Participantes (Presencial)",
                    "Participantes e Status (Híbrido)"
                ]
                
                for i, row_series in df_temp.iterrows():
                    if pd.notna(row_series.iloc[0]) and any(kw in str(row_series.iloc[0]) for kw in header_keywords):
                        if i + 1 < len(df_temp) and "Sessões" in str(df_temp.iloc[i+1, 0]):
                             header_actual_index = i + 2
                        else:
                             header_actual_index = i + 1
                        break
                
                if header_actual_index == -1 or header_actual_index >= len(df_temp):
                    self._log_to_email_tab(f"AVISO: Keyword de cabeçalho não encontrado. Tentando ler com header padrão (primeira linha).")
                    df_report = pd.read_excel(xls, sheet_name=sheet_name_to_read) 
                else:
                    self._log_to_email_tab(f"Keyword de cabeçalho encontrado. Usando linha de índice {header_actual_index} como cabeçalho.")
                    df_report = pd.read_excel(xls, sheet_name=sheet_name_to_read, header=header_actual_index)
                
                df_report.columns = df_report.columns.astype(str).str.strip()
                self._log_to_email_tab(f"Relatório lido da aba: '{sheet_name_to_read}'. Colunas: {list(df_report.columns)}. {len(df_report)} linhas de dados.")
                

            except Exception as e_read:
                raise ValueError(f"Erro ao ler relatório '{os.path.basename(report_path)}': {e_read}")

            required_cols = ["Nome", "Email", "Status"]
            for col in required_cols:
                found_col = find_col_ignore_case(df_report.columns, [col.lower(), col])
                if found_col and found_col != col: df_report.rename(columns={found_col: col}, inplace=True)
                elif col not in df_report.columns: raise ValueError(f"Coluna '{col}' não encontrada no relatório.")
            
            assunto_presentes = self.assunto_presentes_entry.get()
            corpo_presentes = self.corpo_presentes_text.get("1.0", tk.END)
            assunto_faltantes = self.assunto_faltantes_entry.get()
            corpo_faltantes = self.corpo_faltantes_text.get("1.0", tk.END)
            nome_treinamento = self.entries["Nome do treinamento"].get() or "Não Informado"

            total_participantes = len(df_report)
            emails_enviados_count = 0

            for index, row in df_report.iterrows():
                progress_value = (index + 1) / total_participantes
                self.progress_emails.set(progress_value)
                self.root.update_idletasks()

                nome_p = str(row.get("Nome", "")).strip()
                email_p = str(row.get("Email", "")).strip()
                status_p = str(row.get("Status", "")).strip().lower()

                if not email_p or '@' not in email_p:
                    self._log_to_email_tab(f"  Email inválido para {nome_p or 'participante desconhecido'}. Pulando.")
                    continue

                subject, body = None, None
                if status_p == "presente":
                    subject = assunto_presentes.replace("{{nome}}", nome_p).replace("{{nome_treinamento}}", nome_treinamento)
                    body = corpo_presentes.replace("{{nome}}", nome_p).replace("{{nome_treinamento}}", nome_treinamento)
                elif status_p == "falta":
                    subject = assunto_faltantes.replace("{{nome}}", nome_p).replace("{{nome_treinamento}}", nome_treinamento)
                    body = corpo_faltantes.replace("{{nome}}", nome_p).replace("{{nome_treinamento}}", nome_treinamento)
                else:
                    self._log_to_email_tab(f"  Status '{status_p}' não reconhecido para {email_p}. E-mail não enviado.")
                    continue
                
                if self._send_single_outlook_email(email_p, subject, body, attachment_paths=self.anexos_pos_treinamento_paths):
                    emails_enviados_count += 1
                             
            self._log_to_email_tab(f"Processo de envio concluído. {emails_enviados_count} de {total_participantes} e-mails qualificados foram processados para envio.")
            messagebox.showinfo("Envio Concluído", f"{emails_enviados_count} e-mail(s) processado(s). Verifique o log para detalhes.", parent=self.root)

        except ValueError as ve:
            self._log_to_email_tab(f"ERRO: {ve}")
            messagebox.showerror("Erro de Validação", str(ve), parent=self.root)
        except Exception as e:
            error_details = traceback.format_exc()
            self._log_to_email_tab(f"ERRO INESPERADO: {type(e).__name__}: {e}")
            messagebox.showerror("Erro Inesperado no Envio", f"Ocorreu um erro: {e}\nVerifique o log e o console.", parent=self.root)
            print(f"ERRO _enviar_emails_pos_treinamento_async:\n{error_details}")
        finally:
            self.progress_emails.stop()
            self.progress_emails.set(0)
            self.enviar_pos_treinamento_button.configure(state=tk.NORMAL)
            self.root.update_idletasks()

    def _send_single_outlook_email(self, recipient_email, subject, body, attachment_paths=None):
        """Helper function to send a single email via Outlook."""
        if not recipient_email or '@' not in recipient_email or '.' not in recipient_email.split('@')[-1]:
            self._log_to_email_tab(f"  <!> E-mail inválido pulado: {recipient_email}")
            return False
                
        attachment_paths = attachment_paths or []

        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = recipient_email
            mail.Subject = subject
            mail.Body = body 

            for att_path in attachment_paths:
                if att_path and os.path.exists(att_path):
                    mail.Attachments.Add(os.path.abspath(att_path))
                    self._log_to_email_tab(f"    Anexando: {os.path.basename(att_path)} para {recipient_email}")
                elif att_path:
                    self._log_to_email_tab(f"    <!> Anexo não encontrado, pulando: {att_path} para {recipient_email}")
            
            mail.Send()
            self._log_to_email_tab(f"   > E-mail enviado para: {recipient_email} (Assunto: {subject[:30]}...)")
            return True
        except Exception as e:
            error_details = traceback.format_exc()
            self._log_to_email_tab(f"   <!> Falha ao enviar e-mail para {recipient_email}: {type(e).__name__}: {e}")
            print(f"ERRO ENVIO E-MAIL INDIVIDUAL ({recipient_email}):\n{error_details}")
            return False


# --- Ponto de Entrada Principal (__main__) ---
if __name__ == "__main__":
    # Configurações globais do customtkinter
    ctk.set_appearance_mode("Dark")  # Opções: "System", "Dark", "Light"
    ctk.set_default_color_theme("blue") # Opções: "blue", "green", "dark-blue"
    
    try:
        win32.Dispatch('Scripting.FileSystemObject')
    except Exception as e:
        root_temp = ctk.CTk()
        root_temp.withdraw()
        messagebox.showwarning("Aviso Dependência",f"Não foi possível inicializar o componente COM do Windows:\n{e}\n\nO envio de e-mail via Outlook pode não funcionar corretamente.", parent=root_temp)
        print(f"AVISO: Erro ao inicializar COM: {e}. Funcionalidade de e-mail pode ser afetada.")
        root_temp.destroy()

    login_root = ctk.CTk()
    login_app = LoginApp(login_root)
    login_root.mainloop()